# -*- coding: utf-8 -*-
"""
競輪予測AI「パソ子」- 出力・可視化モジュール
============================================
Excel/CSV出力、背景色付け、買い目表示
"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from datetime import datetime

import pandas as pd
import numpy as np
from loguru import logger

try:
    import openpyxl
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel output limited.")

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except ImportError:
    XLSXWRITER_AVAILABLE = False

from .config import OUTPUT_DIR, OutputConfig, ZoneType, config
from .indicators import RacePrediction


class OutputGenerator:
    """
    出力生成クラス
    
    機能:
    - CSV出力
    - Excel出力（背景色付き）
    - 予想リスト生成
    """
    
    def __init__(self, output_config: Optional[OutputConfig] = None):
        self.config = output_config or config.output
        self.output_dir = OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # ゾーン別背景色
        self.zone_colors = {
            ZoneType.GACHI: self.config.color_gachi,
            ZoneType.BLUE: self.config.color_blue,
            ZoneType.TWILIGHT: self.config.color_twilight,
            ZoneType.RED: self.config.color_red,
        }
    
    def export_to_csv(
        self,
        df: pd.DataFrame,
        filename: Optional[str] = None,
    ) -> Path:
        """
        CSV形式で出力
        
        Args:
            df: 出力データ
            filename: ファイル名（省略時は日時）
            
        Returns:
            Path: 出力ファイルパス
        """
        if filename is None:
            filename = f"prediction_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        if not filename.endswith(".csv"):
            filename += ".csv"
        
        output_path = self.output_dir / filename
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
        
        logger.info(f"Exported CSV to {output_path}")
        return output_path
    
    def export_to_excel(
        self,
        df: pd.DataFrame,
        filename: Optional[str] = None,
        with_colors: bool = True,
    ) -> Path:
        """
        Excel形式で出力（背景色付き）
        
        Args:
            df: 出力データ
            filename: ファイル名
            with_colors: 背景色を付けるか
            
        Returns:
            Path: 出力ファイルパス
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, using CSV instead")
            return self.export_to_csv(df, filename)
        
        if filename is None:
            filename = f"prediction_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        output_path = self.output_dir / filename
        
        # Excelファイル作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "予想"
        
        # ヘッダー書き込み
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # データ書き込み
        for row_idx, row_data in enumerate(df.values, 2):
            zone_value = None
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center")
                
                # ゾーン列の値を記録
                if headers[col_idx - 1] == "zone":
                    zone_value = value
            
            # 背景色を適用
            if with_colors and zone_value:
                color = self._get_color_for_zone(zone_value)
                if color:
                    fill = PatternFill(
                        start_color=color.replace("#", ""),
                        end_color=color.replace("#", ""),
                        fill_type="solid"
                    )
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
        
        # 列幅調整
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
        
        wb.save(output_path)
        logger.info(f"Exported Excel to {output_path}")
        return output_path
    
    def _get_color_for_zone(self, zone_str: str) -> Optional[str]:
        """ゾーン文字列から色を取得"""
        for zone_type, color in self.zone_colors.items():
            if zone_type.value == zone_str:
                return color
        return None
    
    def export_predictions(
        self,
        predictions: List[RacePrediction],
        filename: Optional[str] = None,
        format: str = "excel",
    ) -> Path:
        """
        予測結果を出力
        
        Args:
            predictions: 予測結果リスト
            filename: ファイル名
            format: 出力形式（excel, csv）
            
        Returns:
            Path: 出力ファイルパス
        """
        from .indicators import IndicatorCalculator
        
        calculator = IndicatorCalculator()
        df = calculator.summarize_predictions(predictions)
        
        if format == "excel":
            return self.export_to_excel(df, filename)
        else:
            return self.export_to_csv(df, filename)
    
    def generate_race_card(
        self,
        prediction: RacePrediction,
    ) -> str:
        """
        レースカード（予想表）を文字列で生成
        
        Args:
            prediction: 予測結果
            
        Returns:
            str: レースカード文字列
        """
        lines = []
        
        # ヘッダー
        lines.append("=" * 50)
        lines.append(f"{prediction.venue_name} {prediction.race_no}R")
        lines.append(f"日付: {prediction.date}")
        lines.append("=" * 50)
        
        # ゾーン
        lines.append(f"【{prediction.zone.value}】")
        lines.append("")
        
        # 指標
        lines.append(f"A率: {prediction.a_rate:.1%}")
        lines.append(f"CT値: {prediction.ct_value:.1f}")
        lines.append(f"KS値: {prediction.ks_value:.3f}")
        lines.append("")
        
        # ランキング
        lines.append("▼ 予想ランキング")
        lines.append("-" * 30)
        
        rankings = prediction.rankings.head(9)
        for idx, row in rankings.iterrows():
            rank_symbol = row.get("rank_symbol", "-")
            car_num = int(row.get("car_number", 0))
            player_name = row.get("player_name", "")
            proba = row.get("pred_proba", 0)
            
            line = f"{rank_symbol}: {car_num}番 {player_name} ({proba:.1%})"
            lines.append(line)
        
        lines.append("")
        
        # 推奨買い目
        lines.append("▼ 推奨買い目")
        lines.append("-" * 30)
        for bet in prediction.recommended_bets:
            lines.append(f"  {bet}")
        
        lines.append("=" * 50)
        
        return "\n".join(lines)
    
    def generate_daily_report(
        self,
        predictions: List[RacePrediction],
        date: str,
    ) -> str:
        """
        1日分の予想レポートを生成
        
        Args:
            predictions: 予測結果リスト
            date: 日付
            
        Returns:
            str: レポート文字列
        """
        lines = []
        
        # ヘッダー
        lines.append("#" * 60)
        lines.append(f"# 競輪予想AI「パソ子」 予想レポート")
        lines.append(f"# 日付: {date}")
        lines.append("#" * 60)
        lines.append("")
        
        # サマリー
        zone_counts = {}
        for p in predictions:
            zone = p.zone.value
            zone_counts[zone] = zone_counts.get(zone, 0) + 1
        
        lines.append("【本日のゾーン分布】")
        for zone, count in zone_counts.items():
            lines.append(f"  {zone}: {count}レース")
        lines.append("")
        
        # ガチゾーンのレース
        gachi_races = [p for p in predictions if p.zone == ZoneType.GACHI]
        if gachi_races:
            lines.append("=" * 60)
            lines.append("【ガチゾーン】本命決着濃厚！")
            lines.append("=" * 60)
            for p in gachi_races:
                lines.append(self.generate_race_card(p))
                lines.append("")
        
        # ブルーゾーンのレース
        blue_races = [p for p in predictions if p.zone == ZoneType.BLUE]
        if blue_races:
            lines.append("=" * 60)
            lines.append("【ブルーゾーン】本命サイド決着")
            lines.append("=" * 60)
            for p in blue_races:
                lines.append(self.generate_race_card(p))
                lines.append("")
        
        # トワイライトゾーン
        twilight_races = [p for p in predictions if p.zone == ZoneType.TWILIGHT]
        if twilight_races:
            lines.append("=" * 60)
            lines.append("【トワイライトゾーン】中穴注意")
            lines.append("=" * 60)
            for p in twilight_races:
                lines.append(self.generate_race_card(p))
                lines.append("")
        
        # レッドゾーン
        red_races = [p for p in predictions if p.zone == ZoneType.RED]
        if red_races:
            lines.append("=" * 60)
            lines.append("【レッドゾーン】荒れ注意！大穴狙い")
            lines.append("=" * 60)
            for p in red_races:
                lines.append(self.generate_race_card(p))
                lines.append("")
        
        return "\n".join(lines)
    
    def save_report(
        self,
        report: str,
        filename: Optional[str] = None,
    ) -> Path:
        """レポートをテキストファイルで保存"""
        if filename is None:
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        output_path = self.output_dir / filename
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(report)
        
        logger.info(f"Saved report to {output_path}")
        return output_path


class BettingFormatter:
    """
    買い目フォーマッタ
    
    様々な形式の買い目を生成
    """
    
    @staticmethod
    def format_sanrenpuku(cars: List[int]) -> str:
        """3連複フォーマット"""
        if len(cars) < 3:
            return ""
        sorted_cars = sorted(cars[:3])
        return f"{sorted_cars[0]}-{sorted_cars[1]}-{sorted_cars[2]}"
    
    @staticmethod
    def format_sanrentan(first: int, second: int, third: int) -> str:
        """3連単フォーマット"""
        return f"{first}→{second}→{third}"
    
    @staticmethod
    def format_nishafuku(cars: List[int]) -> str:
        """2車複フォーマット"""
        if len(cars) < 2:
            return ""
        sorted_cars = sorted(cars[:2])
        return f"{sorted_cars[0]}-{sorted_cars[1]}"
    
    @staticmethod
    def format_nishatan(first: int, second: int) -> str:
        """2車単フォーマット"""
        return f"{first}→{second}"
    
    @staticmethod
    def format_formation(
        first_box: List[int],
        second_box: List[int],
        third_box: List[int],
    ) -> str:
        """フォーメーションフォーマット"""
        first_str = ",".join(map(str, first_box))
        second_str = ",".join(map(str, second_box))
        third_str = ",".join(map(str, third_box))
        return f"[{first_str}]-[{second_str}]-[{third_str}]"
    
    @staticmethod
    def generate_box_bets(
        cars: List[int],
        bet_type: str = "sanrenpuku",
    ) -> List[str]:
        """
        ボックス買い目を生成
        
        Args:
            cars: 対象車番リスト
            bet_type: 賭式（sanrenpuku, nishafuku）
            
        Returns:
            List[str]: 買い目リスト
        """
        from itertools import combinations
        
        bets = []
        
        if bet_type == "sanrenpuku":
            for combo in combinations(cars, 3):
                bets.append(BettingFormatter.format_sanrenpuku(list(combo)))
        elif bet_type == "nishafuku":
            for combo in combinations(cars, 2):
                bets.append(BettingFormatter.format_nishafuku(list(combo)))
        
        return bets
    
    @staticmethod
    def generate_nagashi_bets(
        axis: int,
        others: List[int],
        bet_type: str = "sanrenpuku",
    ) -> List[str]:
        """
        流し買い目を生成
        
        Args:
            axis: 軸車番
            others: 相手車番リスト
            bet_type: 賭式
            
        Returns:
            List[str]: 買い目リスト
        """
        from itertools import combinations
        
        bets = []
        
        if bet_type == "sanrenpuku":
            for combo in combinations(others, 2):
                cars = [axis] + list(combo)
                bets.append(BettingFormatter.format_sanrenpuku(cars))
        elif bet_type == "nishafuku":
            for other in others:
                bets.append(BettingFormatter.format_nishafuku([axis, other]))
        
        return bets
