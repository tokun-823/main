"""
競輪レースデータ スクレイピング & Excel書き込みモジュール  v2.0

【概要】
Gamboo（https://www.gamboo.jp）を主なターゲットとして、
競輪の出走表・レース結果をスクレイピングし、
Excelファイルへ書き込む処理を提供する。

【デフォルト取得範囲】
  2018年1月1日 〜 2025年12月31日（全期間）

【出力Excelファイル構成】
  scraped_data/
  ├── race_data_YYYYMMDD.xlsx              出走表（1日単位）
  ├── race_result_YYYYMMDD.xlsx            レース結果（1日単位）
  ├── training_data_YYYY.xlsx              学習データ（年単位・推奨）
  ├── training_data_YYYYMM.xlsx            学習データ（月単位）
  ├── training_data_START_END.xlsx         学習データ（任意期間）
  └── .scrape_progress.json                中断・再開用進捗ファイル

【中断・再開】
  取得中にCtrl+Cで中断しても .scrape_progress.json に進捗が保存され、
  次回実行時に続きから再開できる。

【コマンドライン実行例】
  # 全期間（2018〜2025年）を取得
  python scraper.py --mode full

  # 年範囲を指定して取得
  python scraper.py --mode full --start-year 2020 --end-year 2022

  # 1日分の出走表を取得
  python scraper.py --mode card --date 2024-12-01

  # 1日分のレース結果を取得
  python scraper.py --mode result --date 2024-11-30

  # ダミーデータでExcel書き込みをテスト
  python scraper.py --mode test
"""

import time
import re
import os
import sys
import json
import argparse
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple

import requests
import urllib3
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# === 以下の1行を追加 ===
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# =====================================================================
# ロギング設定
# =====================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("scraper.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)


# =====================================================================
# 定数・デフォルト設定
# =====================================================================

# ★ デフォルト取得範囲（2018年〜2025年）
DEFAULT_START_YEAR: int = 2018
DEFAULT_END_YEAR:   int = 2025
DEFAULT_START_DATE: str = f"{DEFAULT_START_YEAR}-01-01"
DEFAULT_END_DATE:   str = f"{DEFAULT_END_YEAR}-12-31"

# リクエスト設定
REQUEST_INTERVAL:     float = 1.5   # 通常リクエスト間隔（秒）
RETRY_INTERVAL:       float = 5.0   # リトライ時の待機（秒）
MAX_RETRIES:          int   = 3     # 最大リトライ回数
REQUEST_TIMEOUT:      int   = 15    # タイムアウト（秒）

# Excel分割方法（"year" or "month"）
SPLIT_BY: str = "year"   # 年単位推奨（月単位にするとファイルが多くなる）

# 進捗ファイル名
PROGRESS_FILE: str = ".scrape_progress.json"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
}

GAMBOO_BASE = "https://www.gamboo.jp"

# バンク周長マスタ（場コード → 周長m）
TRACK_BANK_LENGTH: Dict[str, int] = {
    "11": 400,  # 函館
    "12": 333,  # 青森
    "13": 400,  # いわき平
    "14": 333,  # 弘前
    "21": 400,  # 宮城
    "22": 333,  # 大宮
    "23": 400,  # 西武園
    "24": 400,  # 京王閣
    "25": 335,  # 立川
    "31": 333,  # 松戸
    "32": 400,  # 千葉
    "33": 400,  # 前橋
    "34": 400,  # 取手
    "41": 500,  # 宇都宮
    "42": 400,  # 大宮
    "43": 400,  # 熊谷
    "51": 335,  # 平塚
    "52": 400,  # 小田原
    "53": 335,  # 伊東
    "54": 400,  # 静岡
    "61": 400,  # 名古屋
    "62": 335,  # 岐阜
    "63": 400,  # 大垣
    "64": 400,  # 豊橋
    "71": 400,  # 四日市
    "72": 335,  # 松阪
    "73": 400,  # 奈良
    "74": 400,  # 向日町
    "75": 400,  # 和歌山
    "81": 333,  # 岸和田
    "82": 400,  # 玉野
    "83": 400,  # 広島
    "84": 400,  # 防府
    "91": 400,  # 高知
    "92": 400,  # 小倉
    "93": 400,  # 久留米
    "94": 335,  # 武雄
    "95": 400,  # 佐世保
    "96": 500,  # 大村
}

# 競輪場コード → 場名マスタ（コードのみで名前が取れない場合の補完用）
TRACK_NAME_MASTER: Dict[str, str] = {
    "11": "函館", "12": "青森", "13": "いわき平", "14": "弘前",
    "21": "宮城", "22": "前橋", "23": "西武園", "24": "京王閣", "25": "立川",
    "31": "松戸", "32": "千葉", "33": "前橋", "34": "取手",
    "41": "宇都宮", "42": "大宮", "43": "熊谷",
    "51": "平塚", "52": "小田原", "53": "伊東", "54": "静岡",
    "61": "名古屋", "62": "岐阜", "63": "大垣", "64": "豊橋",
    "71": "四日市", "72": "松阪", "73": "奈良", "74": "向日町", "75": "和歌山",
    "81": "岸和田", "82": "玉野", "83": "広島", "84": "防府",
    "91": "高知", "92": "小倉", "93": "久留米", "94": "武雄",
    "95": "佐世保", "96": "大村",
}


# =====================================================================
# 進捗管理クラス（中断・再開）
# =====================================================================

class ScrapeProgress:
    """
    スクレイピングの進捗を JSON ファイルで管理する。
    途中で中断しても、次回実行時に続きから再開できる。
    """

    def __init__(self, progress_dir: str = "."):
        self.filepath = os.path.join(progress_dir, PROGRESS_FILE)
        self._data: Dict = self._load()

    def _load(self) -> Dict:
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"completed_dates": [], "last_updated": "", "total_rows": 0}

    def _save(self):
        self._data["last_updated"] = datetime.now().isoformat()
        with open(self.filepath, "w", encoding="utf-8") as f:
            json.dump(self._data, f, ensure_ascii=False, indent=2)

    def is_done(self, date_str: str) -> bool:
        """指定日の処理が完了しているか"""
        return date_str in self._data.get("completed_dates", [])

    def mark_done(self, date_str: str, rows_added: int = 0):
        """指定日の処理完了をマーク"""
        if date_str not in self._data["completed_dates"]:
            self._data["completed_dates"].append(date_str)
        self._data["total_rows"] = self._data.get("total_rows", 0) + rows_added
        self._save()

    def completed_count(self) -> int:
        return len(self._data.get("completed_dates", []))

    def total_rows(self) -> int:
        return self._data.get("total_rows", 0)

    def reset(self):
        """進捗をリセット（最初からやり直す）"""
        self._data = {"completed_dates": [], "last_updated": "", "total_rows": 0}
        self._save()
        logger.info("進捗ファイルをリセットしました")

    def summary(self) -> str:
        return (
            f"完了済み日付: {self.completed_count()}日 / "
            f"累計行数: {self.total_rows():,}行 / "
            f"最終更新: {self._data.get('last_updated', 'なし')}"
        )


# =====================================================================
# スクレイパー本体
# =====================================================================

class KeirinScraper:
    """
    競輪レースデータ スクレイパー

    【基本使用方法】
        scraper = KeirinScraper(output_dir="./scraped_data")

        # 1日分の出走表を取得
        df, path = scraper.scrape_and_save_race_card("2024-12-01")

        # 1日分のレース結果を取得
        df, path = scraper.scrape_and_save_race_result("2024-11-30")

        # 学習用データを期間指定で収集（年単位で自動分割保存）
        scraper.collect_training_data("2018-01-01", "2025-12-31")

        # 全期間（デフォルト2018〜2025年）を一括収集
        scraper.collect_full_history()
    """

    def __init__(
        self,
        output_dir: str = "./scraped_data",
        request_interval: float = REQUEST_INTERVAL,
        resume: bool = True,
    ):
        """
        Args:
            output_dir:        出力ディレクトリ
            request_interval:  リクエスト間隔（秒）
            resume:            True=中断前の進捗から再開する
        """
        self.output_dir       = output_dir
        self.request_interval = request_interval
        self.resume           = resume

        os.makedirs(output_dir, exist_ok=True)

        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.session.verify = False  # ← この1行を追加（SSL証明書の検証を無効化）

        self.progress = ScrapeProgress(progress_dir=output_dir)
        self.writer   = KeirinExcelWriter(output_dir=output_dir)

        logger.info(f"KeirinScraper 初期化完了  output_dir={output_dir}  resume={resume}")
        if resume and self.progress.completed_count() > 0:
            logger.info(f"  前回進捗: {self.progress.summary()}")

    # ==================================================================
    # 公開メソッド
    # ==================================================================

    def collect_full_history(
        self,
        start_year: int = DEFAULT_START_YEAR,
        end_year:   int = DEFAULT_END_YEAR,
        reset_progress: bool = False,
    ) -> List[str]:
        """
        ★ メインメソッド：全期間（デフォルト2018〜2025年）のレース結果を収集し
          年単位でExcelへ保存する。

        Args:
            start_year:      開始年（デフォルト 2018）
            end_year:        終了年（デフォルト 2025）
            reset_progress:  True=進捗をリセットして最初から取得

        Returns:
            保存したExcelファイルパスのリスト
        """
        logger.info("=" * 60)
        logger.info(f" 全期間データ収集 開始: {start_year}年 〜 {end_year}年")
        logger.info("=" * 60)

        if reset_progress:
            self.progress.reset()

        saved_paths: List[str] = []

        for year in range(start_year, end_year + 1):
            start_date = f"{year}-01-01"
            end_date   = f"{year}-12-31"

            logger.info(f"\n{'─'*60}")
            logger.info(f"  {year}年 データ収集開始")
            logger.info(f"{'─'*60}")

            df_year, path = self._collect_year(year, start_date, end_date)

            if path:
                saved_paths.append(path)
                logger.info(f"  {year}年 完了: {len(df_year):,}行 → {path}")
            else:
                logger.warning(f"  {year}年 データなし（スキップ）")

        logger.info("\n" + "=" * 60)
        logger.info(f" 全期間収集 完了")
        logger.info(f" 保存ファイル数: {len(saved_paths)}")
        logger.info(f" 総進捗: {self.progress.summary()}")
        logger.info("=" * 60)

        return saved_paths

    def collect_training_data(
        self,
        start_date: str = DEFAULT_START_DATE,
        end_date:   str = DEFAULT_END_DATE,
        split_by:   str = SPLIT_BY,
    ) -> List[Tuple[pd.DataFrame, str]]:
        """
        期間指定で学習用データを収集し、年単位または月単位でExcelへ保存。
        """
        logger.info(f"\n学習用データ収集: {start_date} 〜 {end_date}  (split_by={split_by})")

        start = datetime.strptime(start_date, "%Y-%m-%d")
        end   = datetime.strptime(end_date,   "%Y-%m-%d")

        # 分割キーごとにデータを蓄積
        period_rows: Dict[str, List[Dict]] = {}
        
        total_days = (end - start).days + 1
        date_list = [start + timedelta(days=i) for i in range(total_days)]
        skip_days = 0

        with tqdm(date_list, desc="  データ収集", unit="日", leave=True) as pbar:
            for current in pbar:
                date_str   = current.strftime("%Y-%m-%d")
                period_key = current.strftime("%Y") if split_by == "year" else current.strftime("%Y%m")

                # 済みなら skip
                if self.resume and self.progress.is_done(date_str):
                    skip_days += 1
                    pbar.set_postfix({"skip": skip_days})
                    continue

                day_rows = self._fetch_day_results(date_str)

                if day_rows:
                    period_rows.setdefault(period_key, []).extend(day_rows)
                    self.progress.mark_done(date_str, len(day_rows))
                    
                    # 現在の総行数を計算して表示
                    total_rows_now = sum(len(rows) for rows in period_rows.values())
                    pbar.set_postfix({"skip": skip_days, "rows": total_rows_now})

                time.sleep(self.request_interval)

        # 期間ごとに Excel 保存
        results: List[Tuple[pd.DataFrame, str]] = []
        for period_key, rows in period_rows.items():
            if not rows:
                continue
            df = pd.DataFrame(rows)
            filename = f"training_data_{period_key}.xlsx"
            path = self.writer.write_training_data(df, filename)
            results.append((df, path))

        return results

    def scrape_and_save_race_card(
        self, date_str: str, save: bool = True
    ) -> Tuple[pd.DataFrame, str]:
        """
        指定日の出走表をスクレイピングしてExcelへ保存

        Args:
            date_str: "YYYY-MM-DD"
            save:     True なら自動で Excel 保存

        Returns:
            (DataFrame, 保存ファイルパス)
        """
        logger.info(f"\n出走表スクレイピング: {date_str}")

        race_list = self._fetch_race_list(date_str)
        if not race_list:
            logger.warning("対象レースが見つかりませんでした")
            return pd.DataFrame(), ""

        all_rows: List[Dict] = []
        for info in race_list:
            rows = self._scrape_race_card_one(info)
            all_rows.extend(rows)
            time.sleep(self.request_interval)

        df = pd.DataFrame(all_rows)
        logger.info(f"取得完了: {len(df)}行 / {len(race_list)}レース")

        filepath = ""
        if save and not df.empty:
            date_label = date_str.replace("-", "")
            filename   = f"race_data_{date_label}.xlsx"
            filepath   = self.writer.write_race_card(df, filename)

        return df, filepath

    def scrape_and_save_race_result(
        self, date_str: str, save: bool = True
    ) -> Tuple[pd.DataFrame, str]:
        """
        指定日のレース結果をスクレイピングしてExcelへ保存

        Args:
            date_str: "YYYY-MM-DD"
            save:     True なら自動で Excel 保存

        Returns:
            (DataFrame, 保存ファイルパス)
        """
        logger.info(f"\nレース結果スクレイピング: {date_str}")

        day_rows = self._fetch_day_results(date_str)
        df = pd.DataFrame(day_rows)
        logger.info(f"取得完了: {len(df)}行")

        filepath = ""
        if save and not df.empty:
            date_label = date_str.replace("-", "")
            filename   = f"race_result_{date_label}.xlsx"
            filepath   = self.writer.write_race_result(df, filename)

        return df, filepath

    # ==================================================================
    # 内部メソッド：年単位収集
    # ==================================================================

    def _collect_year(
            self, year: int, start_date: str, end_date: str
        ) -> Tuple[pd.DataFrame, str]:
            """1年分のデータを収集してExcelへ保存"""
            start   = datetime.strptime(start_date, "%Y-%m-%d")
            end     = datetime.strptime(end_date,   "%Y-%m-%d")
            all_rows: List[Dict] = []

            total_days  = (end - start).days + 1
            done_days   = 0
            skip_days   = 0

            # 日付リストを生成して tqdm に渡す
            date_list = [start + timedelta(days=i) for i in range(total_days)]
            
            with tqdm(date_list, desc=f"  {year}年", unit="日", leave=True) as pbar:
                for current in pbar:
                    date_str = current.strftime("%Y-%m-%d")

                    # 中断再開：取得済みの日付はスキップ
                    if self.resume and self.progress.is_done(date_str):
                        skip_days += 1
                        pbar.set_postfix({"skip": skip_days, "rows": len(all_rows)})
                        continue

                    day_rows = self._fetch_day_results(date_str)

                    if day_rows:
                        all_rows.extend(day_rows)
                        self.progress.mark_done(date_str, len(day_rows))
                        done_days += 1
                        pbar.set_postfix({"skip": skip_days, "rows": len(all_rows)})
                    else:
                        # 開催なし・取得失敗の日は完了マークせず（翌日再取得可能）
                        pass

                    time.sleep(self.request_interval)

            df = pd.DataFrame(all_rows)

            if df.empty:
                logger.info(f"  {year}年: 取得データなし（全てスキップ済み or 取得失敗）")
                return df, ""

            filename = f"training_data_{year}.xlsx"
            path = self.writer.write_training_data(df, filename)

            logger.info(
                f"  {year}年完了: {len(df):,}行 / {df['race_id'].nunique():,}レース "
                f"→ 保存: {path}"
            )
            return df, path

    # ==================================================================
    # 内部メソッド：1日分の結果取得
    # ==================================================================

    def _fetch_day_results(self, date_str: str) -> List[Dict]:
        """指定日の全レース結果を取得"""
        race_list = self._fetch_race_list(date_str)
        if not race_list:
            return []

        all_rows: List[Dict] = []
        for info in race_list:
            rows = self._scrape_race_result_one(info)
            all_rows.extend(rows)
            time.sleep(self.request_interval)

        return all_rows

    # ==================================================================
    # 内部メソッド：レース一覧取得
    # ==================================================================

    def _fetch_race_list(self, date_str: str) -> List[Dict]:
        """
        指定日に開催されているレース一覧を取得

        Returns:
            [{"date":..., "track_code":..., "track_name":...,
              "race_number":..., "url":...}, ...]
        """
        date_fmt = date_str.replace("-", "")          # 20241201
        url = f"{GAMBOO_BASE}/keirin/calendar/{date_fmt[:6]}/{date_fmt}/"

        soup = self._get_with_retry(url)
        if soup is None:
            return []

        race_list: List[Dict] = []

        for a_tag in soup.select("a[href*='/keirin/race/']"):
            href = a_tag.get("href", "")
            m = re.search(r"/keirin/race/(\d{8})/(\d{2})(\d{2})/", href)
            if not m:
                continue

            date_code  = m.group(1)
            track_code = m.group(2)
            race_num   = int(m.group(3))

            # 場名取得（リンクテキスト → マスタ補完）
            raw_name   = a_tag.get_text(strip=True)
            track_name = raw_name if raw_name else TRACK_NAME_MASTER.get(track_code, f"場{track_code}")

            race_list.append({
                "date":        f"{date_code[:4]}-{date_code[4:6]}-{date_code[6:]}",
                "track_code":  track_code,
                "track_name":  track_name,
                "race_number": race_num,
                "url":         GAMBOO_BASE + href,
            })

        # 重複除去
        seen   = set()
        unique = []
        for r in race_list:
            key = (r["date"], r["track_code"], r["race_number"])
            if key not in seen:
                seen.add(key)
                unique.append(r)

        if unique:
            logger.debug(f"  {date_str}: {len(unique)}レース")
        return unique

    # ==================================================================
    # 内部メソッド：出走表パース
    # ==================================================================

    def _scrape_race_card_one(self, info: Dict) -> List[Dict]:
        """1レース分の出走表をスクレイピング"""
        soup = self._get_with_retry(info["url"])
        if soup is None:
            return []

        bank_length = TRACK_BANK_LENGTH.get(info["track_code"], 400)
        rows: List[Dict] = []

        table = (
            soup.select_one("table.race-card, table.player-list, table#race-table")
            or soup.find("table")
        )
        if table is None:
            return []

        for tr in table.select("tbody tr, tr.player-row"):
            td_list = tr.find_all("td")
            if len(td_list) < 5:
                continue
            try:
                car_number  = self._to_int(self._td_text(td_list, 0))
                player_name = self._td_text(td_list, 1)
                player_grade = self._td_text(td_list, 2)
                race_score  = self._to_float(self._td_text(td_list, 4))
                back_count  = self._to_int(self._td_text(td_list, 5))

                rows.append({
                    "race_id":        self._make_race_id(info),
                    "race_date":      info["date"],
                    "track_name":     info["track_name"],
                    "track_code":     info["track_code"],
                    "race_number":    info["race_number"],
                    "grade":          self._extract_grade(soup),
                    "bank_length":    bank_length,
                    "is_girls":       self._is_girls(soup),
                    "is_challenge":   self._is_challenge(soup),
                    "car_number":     car_number,
                    "frame_number":   car_number,
                    "player_name":    player_name,
                    "player_grade":   player_grade,
                    "race_score":     race_score,
                    "back_count":     back_count,
                    "line_formation": self._extract_line(soup),
                    "is_line_leader": 0,
                    "accident_flag":  0,
                    "finish_position": None,
                })
            except Exception:
                continue

        return self._assign_line_leader(rows)

    # ==================================================================
    # 内部メソッド：レース結果パース
    # ==================================================================

    def _scrape_race_result_one(self, info: Dict) -> List[Dict]:
        """1レース分のレース結果をスクレイピング（着順を含む）"""
        result_url = info["url"].replace("/race/", "/result/")
        soup = self._get_with_retry(result_url) or self._get_with_retry(info["url"])
        if soup is None:
            return []

        bank_length = TRACK_BANK_LENGTH.get(info["track_code"], 400)
        rows: List[Dict] = []

        table = (
            soup.select_one("table.result-table, table.race-result, table")
        )
        if table is None:
            return []

        finish_map = self._parse_finish_order(soup)

        for tr in table.select("tbody tr, tr"):
            td_list = tr.find_all("td")
            if len(td_list) < 5:
                continue
            try:
                car_number   = self._to_int(self._td_text(td_list, 0))
                player_name  = self._td_text(td_list, 1)
                player_grade = self._td_text(td_list, 2)
                race_score   = self._to_float(self._td_text(td_list, 4))
                back_count   = self._to_int(self._td_text(td_list, 5))

                # 落車・失格チェック
                row_text      = tr.get_text()
                accident_flag = int(any(kw in row_text for kw in ["落車", "失格", "棄権", "妨害"]))

                rows.append({
                    "race_id":        self._make_race_id(info),
                    "race_date":      info["date"],
                    "track_name":     info["track_name"],
                    "track_code":     info["track_code"],
                    "race_number":    info["race_number"],
                    "grade":          self._extract_grade(soup),
                    "bank_length":    bank_length,
                    "is_girls":       self._is_girls(soup),
                    "is_challenge":   self._is_challenge(soup),
                    "car_number":     car_number,
                    "frame_number":   car_number,
                    "player_name":    player_name,
                    "player_grade":   player_grade,
                    "race_score":     race_score,
                    "back_count":     back_count,
                    "line_formation": self._extract_line(soup),
                    "is_line_leader": 0,
                    "accident_flag":  accident_flag,
                    "finish_position": finish_map.get(car_number, 99),
                })
            except Exception:
                continue

        return self._assign_line_leader(rows)

    # ==================================================================
    # HTML解析ヘルパー
    # ==================================================================

    def _get_with_retry(
        self, url: str, retries: int = MAX_RETRIES
    ) -> Optional[BeautifulSoup]:
        """リトライ付き GET リクエスト"""
        for attempt in range(1, retries + 1):
            try:
                resp = self.session.get(url, timeout=REQUEST_TIMEOUT)
                resp.raise_for_status()
                resp.encoding = resp.apparent_encoding
                return BeautifulSoup(resp.text, "lxml")
            except requests.HTTPError as e:
                if resp.status_code == 404:
                    logger.debug(f"  404 Not Found: {url}")
                    return None
                logger.warning(f"  [{attempt}/{retries}] HTTPエラー: {e}  URL: {url}")
            except requests.RequestException as e:
                logger.warning(f"  [{attempt}/{retries}] リクエスト失敗: {e}  URL: {url}")

            if attempt < retries:
                time.sleep(RETRY_INTERVAL)

        logger.error(f"  全リトライ失敗: {url}")
        return None

    def _td_text(self, td_list, idx: int, default: str = "") -> str:
        if idx >= len(td_list):
            return default
        return td_list[idx].get_text(strip=True)

    def _to_float(self, s: str) -> float:
        try:
            return float(re.sub(r"[^\d.]", "", s))
        except Exception:
            return 0.0

    def _to_int(self, s) -> int:
        try:
            return int(re.sub(r"[^\d]", "", str(s)))
        except Exception:
            return 0

    def _make_race_id(self, info: Dict) -> str:
        date = info["date"].replace("-", "")
        return f"{date}{info['track_code']}{info['race_number']:02d}"

    def _extract_grade(self, soup: BeautifulSoup) -> str:
        text = soup.get_text()
        for kw in ["G1", "G2", "G3", "FI", "F1", "F2"]:
            if kw in text:
                return kw
        return "F1"

    def _is_girls(self, soup: BeautifulSoup) -> bool:
        return "ガールズ" in soup.get_text()

    def _is_challenge(self, soup: BeautifulSoup) -> bool:
        return "チャレンジ" in soup.get_text()

    def _extract_line(self, soup: BeautifulSoup) -> str:
        text = soup.get_text()
        m = re.search(r"(\d)-(\d)-?(\d)?", text)
        if m:
            parts = [m.group(1), m.group(2)]
            if m.group(3):
                parts.append(m.group(3))
            return "-".join(parts)
        return ""

    def _parse_finish_order(self, soup: BeautifulSoup) -> Dict[int, int]:
        finish_map: Dict[int, int] = {}
        for tag in soup.select("td.finish-order, td.rank, td"):
            text = tag.get_text(strip=True)
            m = re.match(r"^(\d)着?\s*.*?(\d+)番?$", text)
            if m:
                finish_map[int(m.group(2))] = int(m.group(1))
        return finish_map

    def _assign_line_leader(self, rows: List[Dict]) -> List[Dict]:
        """ライン先頭フラグ推定（車番1,4,7 → 先頭とする簡易実装）"""
        leader_nums = {1, 4, 7}
        for r in rows:
            r["is_line_leader"] = 1 if r.get("car_number") in leader_nums else 0
        return rows


# =====================================================================
# Excel書き込みクラス
# =====================================================================

class KeirinExcelWriter:
    """スクレイピングしたデータをExcelへ書き込む"""

    RACE_CARD_COLUMNS = [
        "race_id", "race_date", "track_name", "track_code",
        "race_number", "grade", "bank_length",
        "is_girls", "is_challenge",
        "car_number", "frame_number", "player_name", "player_grade",
        "race_score", "back_count", "line_formation",
        "is_line_leader", "accident_flag",
    ]
    RESULT_COLUMNS = RACE_CARD_COLUMNS + ["finish_position"]

    COLUMN_LABELS = {
        "race_id":         "レースID",
        "race_date":       "開催日",
        "track_name":      "競輪場",
        "track_code":      "場コード",
        "race_number":     "R番号",
        "grade":           "グレード",
        "bank_length":     "バンク(m)",
        "is_girls":        "ガールズ",
        "is_challenge":    "チャレンジ",
        "car_number":      "車番",
        "frame_number":    "枠番",
        "player_name":     "選手名",
        "player_grade":    "選手級",
        "race_score":      "競走得点",
        "back_count":      "バック回数",
        "line_formation":  "ライン",
        "is_line_leader":  "先頭FLG",
        "accident_flag":   "事故FLG",
        "finish_position": "着順",
    }

    GRADE_COLORS = {
        "G1": "FFFFE0B2",
        "G2": "FFFFF9C4",
        "G3": "FFE8F5E9",
    }

    def __init__(self, output_dir: str = "./scraped_data"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    # ── 出走表 ──────────────────────────────────────────────────────

    def write_race_card(self, df: pd.DataFrame, filename: str) -> str:
        """出走表 DataFrame を Excel へ書き込む"""
        filepath = os.path.join(self.output_dir, filename)
        cols   = [c for c in self.RACE_CARD_COLUMNS if c in df.columns]
        df_out = df[cols].rename(columns=self.COLUMN_LABELS)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="出走表")
            for track, grp in df.groupby("track_name"):
                g_out = grp[[c for c in self.RACE_CARD_COLUMNS if c in grp.columns]]
                g_out = g_out.rename(columns=self.COLUMN_LABELS)
                g_out.to_excel(writer, index=False, sheet_name=str(track)[:31])

        self._apply_card_format(filepath)
        logger.info(f"  [保存] 出走表: {filepath}")
        return filepath

    # ── レース結果 ──────────────────────────────────────────────────

    def write_race_result(self, df: pd.DataFrame, filename: str) -> str:
        """レース結果 DataFrame を Excel へ書き込む"""
        filepath = os.path.join(self.output_dir, filename)
        cols   = [c for c in self.RESULT_COLUMNS if c in df.columns]
        df_out = df[cols].rename(columns=self.COLUMN_LABELS)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="レース結果")
            if "finish_position" in df.columns:
                self._make_result_stats(df).to_excel(
                    writer, index=False, sheet_name="統計"
                )

        self._apply_result_format(filepath)
        logger.info(f"  [保存] 結果: {filepath}")
        return filepath

    # ── 学習データ ──────────────────────────────────────────────────

    def write_training_data(self, df: pd.DataFrame, filename: str) -> str:
        """学習用 DataFrame を Excel へ書き込む（グレード別シート付き）"""
        filepath = os.path.join(self.output_dir, filename)
        cols   = [c for c in self.RESULT_COLUMNS if c in df.columns]
        df_out = df[cols].rename(columns=self.COLUMN_LABELS)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df_out.to_excel(writer, index=False, sheet_name="学習データ(全)")

            if "grade" in df.columns:
                for grade, grp in df.groupby("grade"):
                    g_cols = [c for c in self.RESULT_COLUMNS if c in grp.columns]
                    grp[g_cols].rename(columns=self.COLUMN_LABELS).to_excel(
                        writer, index=False, sheet_name=f"grade_{grade}"[:31]
                    )

            self._make_training_summary(df).to_excel(
                writer, index=False, sheet_name="サマリー"
            )

        self._apply_training_format(filepath)
        logger.info(
            f"  [保存] 学習データ: {filepath}  "
            f"({len(df):,}行 / {df['race_id'].nunique():,}レース)"
        )
        return filepath

    # ── 書式適用 ────────────────────────────────────────────────────

    def _apply_card_format(self, filepath: str):
        wb = load_workbook(filepath)
        for ws in wb.worksheets:
            self._fmt_header(ws)
            self._auto_width(ws)
            ws.freeze_panes = "A2"
            headers   = [c.value for c in ws[1]]
            grade_col = self._find_col(headers, "グレード")
            if grade_col:
                for ri in range(2, ws.max_row + 1):
                    grade = str(ws.cell(ri, grade_col).value or "")
                    color = self.GRADE_COLORS.get(grade, "FFFFFFFF")
                    fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for ci in range(1, ws.max_column + 1):
                        ws.cell(ri, ci).fill = fill
        wb.save(filepath)

    def _apply_result_format(self, filepath: str):
        wb = load_workbook(filepath)
        for ws in wb.worksheets:
            self._fmt_header(ws)
            self._auto_width(ws)
            ws.freeze_panes = "A2"
            headers    = [c.value for c in ws[1]]
            finish_col = self._find_col(headers, "着順")
            acc_col    = self._find_col(headers, "事故FLG")
            top3_fill  = {1: "FFFFFFCC", 2: "FFE3F2FD", 3: "FFE8F5E9"}
            if finish_col:
                for ri in range(2, ws.max_row + 1):
                    val = ws.cell(ri, finish_col).value
                    if isinstance(val, (int, float)) and int(val) in top3_fill:
                        fill = PatternFill(
                            start_color=top3_fill[int(val)],
                            end_color=top3_fill[int(val)],
                            fill_type="solid",
                        )
                        for ci in range(1, ws.max_column + 1):
                            ws.cell(ri, ci).fill = fill
                        ws.cell(ri, finish_col).font = Font(bold=True)
            if acc_col:
                for ri in range(2, ws.max_row + 1):
                    if ws.cell(ri, acc_col).value == 1:
                        fill = PatternFill(start_color="FFFFCCBC", end_color="FFFFCCBC", fill_type="solid")
                        for ci in range(1, ws.max_column + 1):
                            ws.cell(ri, ci).fill = fill
        wb.save(filepath)

    def _apply_training_format(self, filepath: str):
        wb = load_workbook(filepath)
        for ws in wb.worksheets:
            self._fmt_header(ws)
            self._auto_width(ws)
            ws.freeze_panes = "A2"
        wb.save(filepath)

    def _fmt_header(self, ws):
        fill   = PatternFill(start_color="FF1565C0", end_color="FF1565C0", fill_type="solid")
        font   = Font(color="FFFFFFFF", bold=True, size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font; cell.alignment = center
        ws.row_dimensions[1].height = 25

    def _auto_width(self, ws, min_w: int = 8, max_w: int = 30):
        for col_cells in ws.columns:
            letter  = get_column_letter(col_cells[0].column)
            max_len = max(
                (sum(2 if ord(c) > 127 else 1 for c in str(cell.value or ""))
                 for cell in col_cells),
                default=0,
            )
            ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))

    def _find_col(self, headers: List, label: str) -> Optional[int]:
        try:
            return headers.index(label) + 1
        except ValueError:
            return None

    # ── 統計シート生成 ──────────────────────────────────────────────

    def _make_result_stats(self, df: pd.DataFrame) -> pd.DataFrame:
        rows = []
        for race_id, g in df.groupby("race_id"):
            rows.append({
                "レースID":  race_id,
                "開催日":    g["race_date"].iloc[0],
                "競輪場":    g["track_name"].iloc[0],
                "R番号":     g["race_number"].iloc[0],
                "グレード":  g["grade"].iloc[0],
                "参加人数":  len(g),
                "1着車番":   g.loc[g["finish_position"] == 1, "car_number"].values[0]
                             if (g["finish_position"] == 1).any() else "",
                "2着車番":   g.loc[g["finish_position"] == 2, "car_number"].values[0]
                             if (g["finish_position"] == 2).any() else "",
                "3着車番":   g.loc[g["finish_position"] == 3, "car_number"].values[0]
                             if (g["finish_position"] == 3).any() else "",
                "事故有":    int(g["accident_flag"].sum()),
            })
        return pd.DataFrame(rows)

    def _make_training_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        rows = [
            {"項目": "総行数",       "値": len(df)},
            {"項目": "総レース数",   "値": df["race_id"].nunique()},
            {"項目": "開始日",       "値": df["race_date"].min()},
            {"項目": "終了日",       "値": df["race_date"].max()},
            {"項目": "競輪場数",     "値": df["track_name"].nunique()},
            {"項目": "3着以内(1)",   "値": int((df.get("finish_position", pd.Series()) <= 3).sum())},
            {"項目": "落車・失格数", "値": int(df.get("accident_flag", pd.Series(0)).sum())},
        ]
        if "grade" in df.columns:
            for grade, g in df.groupby("grade"):
                rows.append({"項目": f"グレード{grade}レース数", "値": g["race_id"].nunique()})
        return pd.DataFrame(rows)


# =====================================================================
# 統合パイプライン（パソ子AI の main.py から呼ばれる）
# =====================================================================

class KeirinDataPipeline:
    """
    スクレイピング → Excel書き込み を一気通貫で実行するパイプライン。
    パソ子AI（main.py）の PasukoAI から呼び出すことを想定。

    使用方法:
        pipeline = KeirinDataPipeline(output_dir="./scraped_data")

        # ★ 2018〜2025年の全データを取得（年単位でExcel保存・中断再開対応）
        paths = pipeline.run_full_history()

        # 任意の年範囲
        paths = pipeline.run_full_history(start_year=2020, end_year=2022)

        # 出走表（当日）
        df, path = pipeline.run_race_card("2024-12-01")

        # レース結果（昨日）
        df, path = pipeline.run_race_result("2024-11-30")

        # 学習用データ（任意期間・年単位分割）
        results = pipeline.run_training_data("2022-01-01", "2022-12-31")

        # 保存したExcelをパソ子AIに渡す形式で読み込む
        df = pipeline.load_excel_for_pasuko("./scraped_data/training_data_2022.xlsx")
    """

    def __init__(
        self,
        output_dir: str = "./scraped_data",
        request_interval: float = REQUEST_INTERVAL,
        resume: bool = True,
    ):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.scraper = KeirinScraper(
            output_dir=output_dir,
            request_interval=request_interval,
            resume=resume,
        )

    # ──────────────────────────────────────────────────────────────

    def run_full_history(
        self,
        start_year: int = DEFAULT_START_YEAR,
        end_year:   int = DEFAULT_END_YEAR,
        reset_progress: bool = False,
    ) -> List[str]:
        """
        ★ 2018〜2025年（デフォルト）の全レース結果を取得し
          年単位のExcelファイルに保存する。

        Args:
            start_year:     開始年（デフォルト 2018）
            end_year:       終了年（デフォルト 2025）
            reset_progress: True=最初から取得し直す

        Returns:
            保存したExcelファイルパスのリスト
        """
        return self.scraper.collect_full_history(
            start_year=start_year,
            end_year=end_year,
            reset_progress=reset_progress,
        )

    def run_race_card(self, date_str: str) -> Tuple[pd.DataFrame, str]:
        """出走表をスクレイピング → Excel保存"""
        return self.scraper.scrape_and_save_race_card(date_str, save=True)

    def run_race_result(self, date_str: str) -> Tuple[pd.DataFrame, str]:
        """レース結果をスクレイピング → Excel保存"""
        return self.scraper.scrape_and_save_race_result(date_str, save=True)

    def run_training_data(
        self,
        start_date: str = DEFAULT_START_DATE,
        end_date:   str = DEFAULT_END_DATE,
        split_by:   str = SPLIT_BY,
    ) -> List[Tuple[pd.DataFrame, str]]:
        """期間指定でデータ収集 → 年/月単位でExcel保存"""
        return self.scraper.collect_training_data(
            start_date=start_date,
            end_date=end_date,
            split_by=split_by,
        )

    def load_excel_for_pasuko(self, filepath: str) -> pd.DataFrame:
        """
        保存したExcelをパソ子AI（main.py）に渡せる形式のDataFrameに変換して返す

        Args:
            filepath: 保存済みExcelのパス

        Returns:
            英語列名のDataFrame（PasukoAI.run_full_pipeline に渡せる形式）
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"ファイルが見つかりません: {filepath}")

        df = pd.read_excel(filepath, sheet_name=0)
        inv = {v: k for k, v in KeirinExcelWriter.COLUMN_LABELS.items()}
        df  = df.rename(columns=inv)
        logger.info(f"Excelロード完了: {len(df):,}行  {filepath}")
        return df

    def show_progress(self):
        """現在の進捗状況を表示"""
        print(self.scraper.progress.summary())


# =====================================================================
# コマンドライン インターフェース
# =====================================================================

def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="競輪レースデータ スクレイパー",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  # 2018〜2025年 全期間を取得（推奨: 中断再開対応）
  python scraper.py --mode full

  # 年範囲を指定
  python scraper.py --mode full --start-year 2020 --end-year 2022

  # 進捗をリセットして最初から
  python scraper.py --mode full --reset

  # 1日分の出走表
  python scraper.py --mode card --date 2024-12-01

  # 1日分のレース結果
  python scraper.py --mode result --date 2024-11-30

  # ダミーデータでExcel書き込みテスト
  python scraper.py --mode test

  # 進捗確認のみ
  python scraper.py --mode progress
        """,
    )
    p.add_argument(
        "--mode",
        choices=["full", "card", "result", "test", "progress"],
        default="test",
        help=(
            "full=全期間取得, card=出走表1日, result=結果1日, "
            "test=ダミーテスト, progress=進捗確認"
        ),
    )
    p.add_argument(
        "--start-year", type=int, default=DEFAULT_START_YEAR,
        help=f"取得開始年（デフォルト: {DEFAULT_START_YEAR}）"
    )
    p.add_argument(
        "--end-year", type=int, default=DEFAULT_END_YEAR,
        help=f"取得終了年（デフォルト: {DEFAULT_END_YEAR}）"
    )
    p.add_argument(
        "--date", type=str, default=datetime.today().strftime("%Y-%m-%d"),
        help="対象日（YYYY-MM-DD）。--mode card/result 時に使用"
    )
    p.add_argument(
        "--output-dir", type=str, default="./scraped_data",
        help="出力ディレクトリ（デフォルト: ./scraped_data）"
    )
    p.add_argument(
        "--interval", type=float, default=REQUEST_INTERVAL,
        help=f"リクエスト間隔（秒）（デフォルト: {REQUEST_INTERVAL}）"
    )
    p.add_argument(
        "--reset", action="store_true",
        help="進捗をリセットして最初から取得する"
    )
    p.add_argument(
        "--no-resume", action="store_true",
        help="中断再開を無効にする（毎回最初から）"
    )
    return p


def _run_test(output_dir: str):
    """ダミーデータを使った Excel 書き込みテスト"""
    import numpy as np

    logger.info("=" * 60)
    logger.info(" ダミーデータ Excel 書き込みテスト")
    logger.info("=" * 60)

    np.random.seed(42)
    n_races, n_players = 10, 9
    rows = []
    for r in range(1, n_races + 1):
        race_id = f"2024120151{r:02d}"
        perm    = np.random.permutation(n_players) + 1
        for car in range(1, n_players + 1):
            rows.append({
                "race_id":        race_id,
                "race_date":      "2024-12-01",
                "track_name":     "平塚",
                "track_code":     "51",
                "race_number":    r,
                "grade":          np.random.choice(["G3", "F1", "F2"]),
                "bank_length":    335,
                "is_girls":       False,
                "is_challenge":   False,
                "car_number":     car,
                "frame_number":   car,
                "player_name":    f"選手{car:02d}",
                "player_grade":   np.random.choice(["S1", "S2", "A1", "A2"]),
                "race_score":     round(float(np.random.uniform(45, 68)), 2),
                "back_count":     int(np.random.randint(0, 5)),
                "line_formation": "3-3-3",
                "is_line_leader": 1 if car in [1, 4, 7] else 0,
                "accident_flag":  0,
                "finish_position": int(perm[car - 1]),
            })

    df      = pd.DataFrame(rows)
    writer  = KeirinExcelWriter(output_dir=output_dir)
    paths   = [
        writer.write_race_card(df,      "race_data_20241201.xlsx"),
        writer.write_race_result(df,    "race_result_20241201.xlsx"),
        writer.write_training_data(df,  "training_data_20241201_20241201.xlsx"),
    ]

    logger.info("\n=== 生成ファイル ===")
    for p in paths:
        size = os.path.getsize(p) / 1024
        logger.info(f"  {os.path.basename(p):<52}  {size:.1f} KB")
    logger.info("テスト完了!")


def main():
    parser = _build_parser()
    args   = parser.parse_args()

    pipeline = KeirinDataPipeline(
        output_dir=args.output_dir,
        request_interval=args.interval,
        resume=not args.no_resume,
    )

    if args.mode == "full":
        logger.info(
            f"全期間取得: {args.start_year}年 〜 {args.end_year}年  "
            f"(resume={'OFF' if args.no_resume else 'ON'}, "
            f"reset={args.reset})"
        )
        paths = pipeline.run_full_history(
            start_year=args.start_year,
            end_year=args.end_year,
            reset_progress=args.reset,
        )
        logger.info(f"\n保存完了: {len(paths)}ファイル")
        for p in paths:
            logger.info(f"  {p}")

    elif args.mode == "card":
        df, path = pipeline.run_race_card(args.date)
        if path:
            logger.info(f"保存: {path}  ({len(df)}行)")

    elif args.mode == "result":
        df, path = pipeline.run_race_result(args.date)
        if path:
            logger.info(f"保存: {path}  ({len(df)}行)")

    elif args.mode == "test":
        _run_test(args.output_dir)

    elif args.mode == "progress":
        pipeline.show_progress()


if __name__ == "__main__":
    main()
