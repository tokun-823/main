"""
データ収集・前処理モジュール  v3.0
競輪レースデータのスクレイピング、クリーニング、特徴量生成を行う

【スクレイピング仕様】
  取得先URL : https://keirin.kdreams.jp/gamboo/keirin-kaisai/race-card/result/
              {YYYYMMDD+場コード}/{YYYYMMDD+場コード}/{R番号}
  取得範囲  : 2018年1月1日 〜 2025年12月31日（デフォルト）
  ID形式    : YYYYMMDD + 場コード(2桁) + レース番号(2桁) = 計12桁
              例: 202412015101  →  2024-12-01, 平塚(51), 1R

【race_data_scrape.py との互換性】
  - 同一URLスキームを使用
  - 同一テーブル解析ロジック（p_htm[0], p_htm[6], p_htm[7] or p_htm[3], p_htm[4]）
  - 同一カラム名・ライン構成解析

【出力DataFrame カラム】
  race_id, race_date, track_name, track_code, race_number,
  grade, bank_length, is_girls, is_challenge,
  car_number, frame_number, player_name, player_age, player_class,
  player_grade, race_score, back_count,
  line_formation, line_code, line_position,
  is_line_leader, finish_position, accident_flag,
  race_title, race_start_time, weather, wind_speed
"""

# ===========================================================
# 標準ライブラリ
# ===========================================================
import os
import re
import sys
import json
import time
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any

# ===========================================================
# サードパーティ
# ===========================================================
import numpy as np
import pandas as pd

try:
    import requests
    from bs4 import BeautifulSoup
    _SCRAPING_AVAILABLE = True
except ImportError:
    _SCRAPING_AVAILABLE = False

# ===========================================================
# ロギング
# ===========================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)


# ===========================================================
# 定数
# ===========================================================

# ★ スクレイピング対象URL（race_data_scrape.py と同一）
KDREAMS_RESULT_BASE = (
    "https://keirin.kdreams.jp/gamboo/keirin-kaisai/race-card/result"
)

# Gamboo カレンダー（開催日・場・レース番号の一覧取得用）
GAMBOO_CALENDAR_BASE = "https://www.gamboo.jp/keirin/calendar"
GAMBOO_RACE_BASE     = "https://www.gamboo.jp"

# デフォルト取得範囲
DEFAULT_START_YEAR = 2018
DEFAULT_END_YEAR   = 2025

# リクエスト制御
REQUEST_INTERVAL = 1.5   # 秒
RETRY_INTERVAL   = 5.0   # 秒
MAX_RETRIES      = 3
REQUEST_TIMEOUT  = 15    # 秒

# 進捗ファイル
PROGRESS_FILE = ".dp_scrape_progress.json"

# ★ race_data_scrape.py と同じカラムセット
RACECARD_COLS = [
    "予 想", "好 気 合", "総 評", "枠 番", "車 番",
    "選手名 府県/年齢/期別", "級 班", "脚 質", "ギヤ 倍数", "競走得点",
    "S", "B", "逃", "捲", "差", "マ",
    "1 着", "2 着", "3 着", "着 外",
    "勝 率", "2連 対率", "3連 対率", "level_1", "level_2",
]

GRADE_LIST = ["ＧＰ", "Ｇ１", "Ｇ２", "Ｇ３", "Ｆ１", "Ｆ２"]

# バンク周長マスタ（場コード → m）
TRACK_BANK_LENGTH: Dict[str, int] = {
    "11": 400, "12": 333, "13": 400, "14": 333,
    "21": 400, "22": 333, "23": 400, "24": 400, "25": 335,
    "31": 333, "32": 400, "33": 400, "34": 400,
    "41": 500, "42": 400, "43": 400,
    "51": 335, "52": 400, "53": 335, "54": 400,
    "61": 400, "62": 335, "63": 400, "64": 400,
    "71": 400, "72": 335, "73": 400, "74": 400, "75": 400,
    "81": 333, "82": 400, "83": 400, "84": 400,
    "91": 400, "92": 400, "93": 400, "94": 335, "95": 400, "96": 500,
}

# 場コード → 場名
TRACK_NAME_MASTER: Dict[str, str] = {
    "11": "函館",   "12": "青森",   "13": "いわき平", "14": "弘前",
    "21": "宮城",   "22": "前橋",   "23": "西武園",   "24": "京王閣", "25": "立川",
    "31": "松戸",   "32": "千葉",   "33": "前橋",     "34": "取手",
    "41": "宇都宮", "42": "大宮",   "43": "熊谷",
    "51": "平塚",   "52": "小田原", "53": "伊東",     "54": "静岡",
    "61": "名古屋", "62": "岐阜",   "63": "大垣",     "64": "豊橋",
    "71": "四日市", "72": "松阪",   "73": "奈良",     "74": "向日町", "75": "和歌山",
    "81": "岸和田", "82": "玉野",   "83": "広島",     "84": "防府",
    "91": "高知",   "92": "小倉",   "93": "久留米",   "94": "武雄",
    "95": "佐世保", "96": "大村",
}

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.7,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
}


# ===========================================================
# 進捗管理（中断・再開）
# ===========================================================

class _ScrapeProgress:
    """日単位でスクレイピング進捗を管理する軽量クラス"""

    def __init__(self, output_dir: str = "."):
        self.path = os.path.join(output_dir, PROGRESS_FILE)
        self._data = self._load()

    def _load(self) -> dict:
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"done": [], "total_rows": 0}

    def _save(self):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self._data, f, ensure_ascii=False, indent=2)

    def is_done(self, date_str: str) -> bool:
        return date_str in self._data["done"]

    def mark_done(self, date_str: str, n_rows: int = 0):
        if date_str not in self._data["done"]:
            self._data["done"].append(date_str)
        self._data["total_rows"] += n_rows
        self._save()

    def reset(self):
        self._data = {"done": [], "total_rows": 0}
        self._save()

    @property
    def total_rows(self) -> int:
        return self._data["total_rows"]

    @property
    def done_count(self) -> int:
        return len(self._data["done"])


# ===========================================================
# KeirinDataCollector  ★ スクレイピング実装
# ===========================================================

class KeirinDataCollector:
    """
    競輪データ収集クラス  v3.0

    【主要メソッド】
      scrape_all_races()     : 2018〜2025年の全レースを収集（メイン）
      scrape_races_by_ids()  : race_data_scrape.py 互換の一括取得
      scrape_one_race()      : 1レース分を取得
      get_race_ids_for_date(): 指定日の全レースIDを取得
      load_from_excel()      : 保存済みExcelを読み込み

    【race_data_scrape.py との互換性】
      URLスキーム・テーブル解析・ライン構成解析はすべて同一仕様
    """

    def __init__(
        self,
        start_year: int = DEFAULT_START_YEAR,
        end_year:   int = DEFAULT_END_YEAR,
        output_dir: str = "./scraped_data",
        request_interval: float = REQUEST_INTERVAL,
        resume: bool = True,
    ):
        """
        Args:
            start_year:        取得開始年（デフォルト 2018）
            end_year:          取得終了年（デフォルト 2025）
            output_dir:        Excel保存先ディレクトリ
            request_interval:  リクエスト間隔（秒）
            resume:            True = 中断前の進捗から再開
        """
        self.start_year       = start_year
        self.end_year         = end_year
        self.output_dir       = output_dir
        self.request_interval = request_interval
        self.resume           = resume

        os.makedirs(output_dir, exist_ok=True)

        if _SCRAPING_AVAILABLE:
            self._session = requests.Session()
            self._session.headers.update(HTTP_HEADERS)
        else:
            self._session = None
            logger.warning(
                "requests / beautifulsoup4 がインストールされていません。"
                "スクレイピング機能は使用できません。"
                "pip install requests beautifulsoup4 lxml を実行してください。"
            )

        self._progress = _ScrapeProgress(output_dir)

        if resume and self._progress.done_count > 0:
            logger.info(
                f"前回進捗: 完了済み {self._progress.done_count}日 / "
                f"累計 {self._progress.total_rows:,}行"
            )

    # ──────────────────────────────────────────────────────
    #  ★★★  メイン公開API  ★★★
    # ──────────────────────────────────────────────────────

    def scrape_all_races(
        self,
        start_year: int = None,
        end_year:   int = None,
        save_excel: bool = True,
        reset_progress: bool = False,
    ) -> pd.DataFrame:
        """
        ★ メインメソッド: 2018〜2025年の全レースデータを収集

        処理フロー:
          1. Gamboo カレンダーから日別レース一覧を取得
          2. 各レースIDに対して kdreams.jp から出走表＋結果を取得
          3. ライン構成・レース情報を付加
          4. 年単位でExcelに保存（中断再開対応）

        Args:
            start_year:      開始年（None = self.start_year = 2018）
            end_year:        終了年（None = self.end_year   = 2025）
            save_excel:      True = 年単位でExcelに自動保存
            reset_progress:  True = 進捗リセットして最初から

        Returns:
            全期間の DataFrame（メモリに保持）
        """
        _start = start_year or self.start_year
        _end   = end_year   or self.end_year

        logger.info("=" * 60)
        logger.info(f" 全期間スクレイピング: {_start}年 〜 {_end}年")
        logger.info("=" * 60)

        if reset_progress:
            self._progress.reset()

        all_frames: List[pd.DataFrame] = []

        for year in range(_start, _end + 1):
            logger.info(f"\n{'─'*60}")
            logger.info(f"  {year}年 開始")
            logger.info(f"{'─'*60}")

            df_year = self._collect_year(year, save_excel=save_excel)
            if not df_year.empty:
                all_frames.append(df_year)
                logger.info(f"  {year}年 完了: {len(df_year):,}行")
            else:
                logger.warning(f"  {year}年: データなし")

        if not all_frames:
            logger.warning("取得データがありません")
            return pd.DataFrame()

        df_all = pd.concat(all_frames, ignore_index=True)
        logger.info(
            f"\n全期間完了: {len(df_all):,}行 / "
            f"{df_all['race_id'].nunique():,}レース"
        )
        return df_all

    def scrape_races_by_ids(
        self,
        race_ids: List[str],
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        ★ race_data_scrape.py 互換メソッド

        Args:
            race_ids: 12桁のレースIDリスト
                      例: ["202412015101", "202412015102", ...]

        Returns:
            (info_table, entry_table, return_table)
              - info_table  : レース情報（1行/レース）
              - entry_table : 出走表＋結果（N行/レース）
              - return_table: 払い戻し表
        """
        self._check_scraping()

        info_dict:   Dict[str, Any]        = {}
        entry_dict:  Dict[str, pd.DataFrame] = {}
        return_dict: Dict[str, pd.DataFrame] = {}

        for i, race_id in enumerate(race_ids):
            logger.info(f"[{i+1}/{len(race_ids)}] {race_id}")
            try:
                info, entry, ret = self._scrape_one_race_full(race_id)
                if entry is not None and not entry.empty:
                    info_dict[race_id]   = info
                    entry_dict[race_id]  = entry
                    return_dict[race_id] = ret if ret is not None else pd.DataFrame()
            except Exception as e:
                logger.warning(f"  スキップ ({race_id}): {e}")
            time.sleep(self.request_interval)

        if not entry_dict:
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

        info_table   = pd.DataFrame(info_dict).T
        entry_table  = pd.concat(list(entry_dict.values()),  ignore_index=False)
        return_table = pd.concat(list(return_dict.values()), ignore_index=False)

        return info_table, entry_table, return_table

    def scrape_one_race(
        self,
        race_id: str,
    ) -> pd.DataFrame:
        """
        1レース分のデータを取得して正規化DataFrameで返す

        Args:
            race_id: 12桁レースID (例: "202412015101")

        Returns:
            正規化済みDataFrame（1レース分の出走者行）
        """
        self._check_scraping()
        _, entry, _ = self._scrape_one_race_full(race_id)
        if entry is None or entry.empty:
            return pd.DataFrame()
        return self._normalize_entry(entry, race_id)

    def get_race_ids_for_date(
        self,
        date_str: str,
    ) -> List[str]:
        """
        指定日に開催されているレースのIDリストを返す

        Args:
            date_str: "YYYY-MM-DD"

        Returns:
            12桁レースIDのリスト
            例: ["202412015101", "202412015102", "202412016201", ...]
        """
        self._check_scraping()
        race_list = self._fetch_race_list_gamboo(date_str)
        return [item["race_id"] for item in race_list]

    def scrape_race_data(self, url: str = "") -> pd.DataFrame:
        """
        後方互換メソッド（url は未使用、全期間スクレイピングを呼び出す）

        ※ 通常は scrape_all_races() を直接呼び出してください。
        """
        logger.info("scrape_race_data() → scrape_all_races() を呼び出します")
        return self.scrape_all_races()

    def load_from_excel(self, file_path: str) -> pd.DataFrame:
        """保存済みExcelファイルをDataFrameとして読み込む"""
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            logger.info(f"Excelロード完了: {len(df):,}行  {file_path}")
            return df
        except Exception as e:
            logger.error(f"Excelロード失敗: {e}")
            return pd.DataFrame()

    # ──────────────────────────────────────────────────────
    #  内部: 年単位収集
    # ──────────────────────────────────────────────────────

    def _collect_year(
        self,
        year: int,
        save_excel: bool = True,
    ) -> pd.DataFrame:
        """1年分のデータを収集して返す（進捗管理付き）"""
        start   = datetime(year, 1,  1)
        end     = datetime(year, 12, 31)
        current = start
        rows: List[pd.DataFrame] = []

        total_days = (end - start).days + 1
        processed  = 0

        while current <= end:
            date_str = current.strftime("%Y-%m-%d")

            if self.resume and self._progress.is_done(date_str):
                current += timedelta(days=1)
                continue

            day_rows = self._collect_day(date_str)
            n        = len(day_rows)

            if n > 0:
                rows.append(day_rows)
                self._progress.mark_done(date_str, n)
                processed += 1

                # 進捗ログ（30日ごと）
                if processed % 30 == 0:
                    elapsed = (current - start).days + 1
                    logger.info(
                        f"  {year}年 {elapsed}/{total_days}日  "
                        f"累計 {sum(len(r) for r in rows):,}行"
                    )
            else:
                # 開催なし・取得失敗日は進捗マークせず（翌日再試行可能）
                pass

            current += timedelta(days=1)
            time.sleep(self.request_interval)

        if not rows:
            return pd.DataFrame()

        df_year = pd.concat(rows, ignore_index=True)

        if save_excel and not df_year.empty:
            path = os.path.join(self.output_dir, f"training_data_{year}.xlsx")
            self._save_excel(df_year, path)
            logger.info(f"  Excel保存: {path}")

        return df_year

    # ──────────────────────────────────────────────────────
    #  内部: 1日分収集
    # ──────────────────────────────────────────────────────

    def _collect_day(self, date_str: str) -> pd.DataFrame:
        """指定日の全レースを収集して正規化DataFrameで返す"""
        race_list = self._fetch_race_list_gamboo(date_str)
        if not race_list:
            return pd.DataFrame()

        day_rows: List[pd.DataFrame] = []
        for info in race_list:
            race_id = info["race_id"]
            try:
                df_race = self.scrape_one_race(race_id)
                if not df_race.empty:
                    day_rows.append(df_race)
            except Exception as e:
                logger.debug(f"  スキップ ({race_id}): {e}")
            time.sleep(self.request_interval)

        if not day_rows:
            return pd.DataFrame()
        return pd.concat(day_rows, ignore_index=True)

    # ──────────────────────────────────────────────────────
    #  内部: Gamboo カレンダーからレース一覧取得
    # ──────────────────────────────────────────────────────

    def _fetch_race_list_gamboo(self, date_str: str) -> List[Dict]:
        """
        Gamboo カレンダーページから指定日の開催レース一覧を取得

        Returns:
            [{"race_id": "202412015101",
              "date": "2024-12-01",
              "track_code": "51",
              "track_name": "平塚",
              "race_number": 1}, ...]
        """
        date_fmt = date_str.replace("-", "")           # "20241201"
        cal_url  = f"{GAMBOO_CALENDAR_BASE}/{date_fmt[:6]}/{date_fmt}/"
        soup     = self._get_soup(cal_url)
        if soup is None:
            return []

        race_list: List[Dict] = []
        seen = set()

        # Gamboo のレースリンクパターン: /keirin/race/YYYYMMDD/TTNN/
        for a in soup.select("a[href*='/keirin/race/']"):
            href = a.get("href", "")
            m    = re.search(r"/keirin/race/(\d{8})/(\d{2})(\d{2})/", href)
            if not m:
                continue

            date_code  = m.group(1)   # "20241201"
            track_code = m.group(2)   # "51"
            race_no_s  = m.group(3)   # "01"
            race_no    = int(race_no_s)

            # race_id = YYYYMMDD + track_code(2) + race_no(2) = 12桁
            race_id = f"{date_code}{track_code}{race_no_s}"

            key = (date_code, track_code, race_no_s)
            if key in seen:
                continue
            seen.add(key)

            track_name = TRACK_NAME_MASTER.get(
                track_code,
                a.get_text(strip=True) or f"場{track_code}",
            )

            race_list.append({
                "race_id":     race_id,
                "date":        f"{date_code[:4]}-{date_code[4:6]}-{date_code[6:]}",
                "track_code":  track_code,
                "track_name":  track_name,
                "race_number": race_no,
            })

        logger.debug(f"  {date_str}: {len(race_list)}レース")
        return race_list

    # ──────────────────────────────────────────────────────
    #  内部: 1レース スクレイピング（race_data_scrape.py 同一ロジック）
    # ──────────────────────────────────────────────────────

    def _scrape_one_race_full(
        self,
        race_id: str,
    ) -> Tuple[Optional[Dict], Optional[pd.DataFrame], Optional[pd.DataFrame]]:
        """
        race_data_scrape.py の scrape ループ本体と同一ロジック

        URL構成（race_data_scrape.py と同一）:
          ur_1 = race_id[:10]   例: "2024120151"
          ur_2 = race_id[:-2]   例: "2024120151"
          ur_3 = race_id[-2:]   例: "01"
          URL  = BASE / ur_1 / ur_2 / ur_3
        """
        ur_1 = race_id[:10]
        ur_2 = race_id[:-2]
        ur_3 = race_id[-2:]
        url  = f"{KDREAMS_RESULT_BASE}/{ur_1}/{ur_2}/{ur_3}"

        html = self._get_raw(url)
        if html is None:
            return None, None, None

        soup = BeautifulSoup(html, "lxml")

        # ── pd.read_html で全テーブル取得（race_data_scrape.py 同一） ──
        try:
            p_htm = pd.read_html(url)
        except Exception:
            return None, None, None

        # ── テーブル振り分け（race_data_scrape.py 同一分岐） ──
        try:
            if len(p_htm) == 8:
                pd_racecard  = pd.DataFrame(p_htm[0])
                pd_raceresult = pd.DataFrame(p_htm[6])
                pd_harai      = pd.DataFrame(p_htm[7])
            else:
                pd_racecard  = pd.DataFrame(p_htm[0])
                pd_raceresult = pd.DataFrame(p_htm[3])
                pd_harai      = pd.DataFrame(p_htm[4])
        except (IndexError, Exception):
            return None, None, None

        # ── カラム名設定（race_data_scrape.py 同一） ──
        try:
            pd_racecard.columns = RACECARD_COLS
            pd_racecard = pd_racecard.drop(
                ["好 気 合", "level_1", "level_2"], axis=1, errors="ignore"
            )
        except Exception:
            return None, None, None

        # ── 結果テーブル結合（race_data_scrape.py 同一） ──
        try:
            pd_raceresult = pd_raceresult.sort_values("車 番")
            drop_cols     = ["着差", "上り", "決ま り手", "S ／ B", "勝敗因"]
            pd_raceresult = pd_raceresult.drop(
                [c for c in drop_cols if c in pd_raceresult.columns], axis=1
            )
            pd_racecard = pd.merge(pd_racecard, pd_raceresult, how="inner", on="車 番")
        except Exception:
            pass   # 結果未掲載レースは出走表のみ継続

        # ── 選手名 府県/年齢/期別 の分解（race_data_scrape.py 同一） ──
        try:
            ages      = []
            gr_classes = []
            names     = []
            for name_str in pd_racecard["選手名 府県/年齢/期別"]:
                if "（欠車）" in name_str:
                    name_str = name_str.replace("（欠車）", "")
                parts = name_str.split("/")
                ages.append(int(parts[1]) if len(parts) > 1 else 0)
                gr_classes.append(int(parts[2]) if len(parts) > 2 else 0)
                nm = name_str.split(" ")
                names.append(nm[0] + " " + nm[1] if len(nm) > 1 else nm[0])

            pd_racecard["選手名"] = names
            pd_racecard["年齢"]   = ages
            pd_racecard["期別"]   = gr_classes
            pd_racecard = pd_racecard.drop(["選手名 府県/年齢/期別"], axis=1,
                                            errors="ignore")

            # カラム並び替え（race_data_scrape.py 同一）
            keep = [
                "予 想", "着 順", "総 評", "枠 番", "車 番",
                "選手名", "競走得点", "年齢", "期別", "級 班", "脚 質", "ギヤ 倍数",
                "S", "B", "逃", "捲", "差", "マ",
                "1 着", "2 着", "3 着", "着 外",
                "勝 率", "2連 対率", "3連 対率",
            ]
            exist = [c for c in keep if c in pd_racecard.columns]
            pd_racecard = pd_racecard.loc[:, exist]
        except Exception:
            pass

        # ── ライン構成解析（race_data_scrape.py 同一） ──
        line_n = "mmmmmm" * 18 + "mmmmmm"
        kousei2 = ""
        make_line = []
        try:
            raw_line = soup.find(
                "div", attrs={"class": "line_position_inner"}
            )
            if raw_line:
                line_n, kousei2, make_line = self._parse_line_position(
                    raw_line.text, len(pd_racecard)
                )
        except Exception:
            pass

        # ── ライン情報を出走表に結合（race_data_scrape.py 同一） ──
        pd_racecard2 = pd_racecard.copy()
        try:
            if make_line:
                pd_make_line = pd.DataFrame(make_line, columns=["車 番", "ライン", "番手"])
                pd_racecard2 = pd.merge(pd_racecard2, pd_make_line,
                                        how="inner", on="車 番")
                toku_jyun    = list(range(1, len(pd_make_line) + 1))
                pd_racecard2 = pd_racecard2.sort_values("競走得点", ascending=False)
                pd_racecard2["得点順位"] = toku_jyun
                pd_racecard2 = pd_racecard2.sort_values("車 番")
        except Exception:
            pass

        # ── レース情報取得（race_data_scrape.py 同一） ──
        info = self._parse_race_info(soup, race_id, pd_racecard, kousei2, line_n)

        pd_racecard2.index = [race_id] * len(pd_racecard2)

        return info, pd_racecard2, pd_harai

    # ──────────────────────────────────────────────────────
    #  内部: ライン構成パース（race_data_scrape.py 同一アルゴリズム）
    # ──────────────────────────────────────────────────────

    def _parse_line_position(
        self,
        raw_text: str,
        num_racers: int,
    ) -> Tuple[str, str, List[List]]:
        """
        race_data_scrape.py の line_position_inner 解析と同一ロジック

        Returns:
            (line_n: str, kousei2: str, make_line: List[[車番, ライン, 番手]])
        """
        p      = raw_text.replace("\n", "P")
        line_n_list = ["mmmmmm"] * 18
        n = 0

        for i in range(len(p)):
            po_1 = p[i - 2] if i >= 2 else ""
            po_2 = p[i - 1] if i >= 1 else ""
            if re.findall(r"\d+", p[i]):
                n += 1
                line_n_list[n] = str(p[i])
                if po_1 == po_2:
                    line_n_list[n] = "mmmmmm"
                    n += 1
                    line_n_list[n] = str(p[i])

        line_n = "".join(line_n_list) + "mmmmmm"

        # ライン構成文字列（例: "3-3-3"）
        line_k  = re.findall(r"\d+", line_n)
        kousei  = sorted([str(len(x)) for x in line_k], reverse=True)
        kousei2 = "-".join(kousei)

        # 番手テーブル（make_line）
        make_line: List[List] = []
        for j in range(1, num_racers + 1):
            for k in range(len(line_n)):
                if str(j) == line_n[k]:
                    prev1 = line_n[k - 1] if k >= 1 else "m"
                    prev2 = line_n[k - 2] if k >= 2 else "m"
                    prev3 = line_n[k - 3] if k >= 3 else "m"
                    prev4 = line_n[k - 4] if k >= 4 else "m"
                    next1 = line_n[k + 1] if k + 1 < len(line_n) else "m"

                    if prev1 == "m":
                        if next1 == "m":
                            make_line.append([j, line_n[k], 0])   # 単騎
                        else:
                            make_line.append([j, line_n[k], 1])   # 先頭
                    else:
                        if prev2 == "m":
                            make_line.append([j, prev1, 2])        # 番手
                        elif prev3 == "m":
                            make_line.append([j, prev2, 3])        # 3番手
                        elif prev4 == "m":
                            make_line.append([j, prev3, 4])        # 4番手
                        else:
                            if k >= 5 and line_n[k - 5] == "m":
                                make_line.append([j, prev4, 5])    # 5番手
                    break

        return line_n, kousei2, make_line

    # ──────────────────────────────────────────────────────
    #  内部: レース情報パース（race_data_scrape.py 同一）
    # ──────────────────────────────────────────────────────

    def _parse_race_info(
        self,
        soup: "BeautifulSoup",
        race_id: str,
        pd_racecard: pd.DataFrame,
        kousei2: str,
        line_n: str,
    ) -> Dict:
        """race_data_scrape.py の info_table 構築と同一ロジック"""
        info: Dict[str, Any] = {}

        try:
            race_header       = soup.find("div", attrs={"class": "race_header"}).find("span").text
            race_info_header  = re.findall(r"\w+", race_header)
            race_info_day     = (
                race_id[2:6] + "-" +
                race_info_header[2][:2] + "-" +
                race_info_header[2][3:5]
            ) if len(race_info_header) > 2 else ""

            race_title   = soup.find("div", attrs={"class": "race_title_header"}).find("span").text
            race_stadium = soup.find("span", attrs={"class": "velodrome"}).text
            race_name    = soup.find("span", attrs={"class": "race"}).text
            race_name    = "".join(race_name.split("\u3000"))

            h1_text = soup.find("h1", attrs={"class": "section_title"})
            race_gr = next(
                (t for t in GRADE_LIST if t in h1_text.text),
                "Ｆ２"
            ) if h1_text else "Ｆ２"

            race_start_time = soup.find("dl", attrs={"class": "time"}).find_all("dd")[0].text
            race_condition  = soup.find("p", attrs={"class": "weather_info"}).find_all("span")
            weather         = race_condition[0].text[-1]
            wind_speed      = race_condition[1].text[2:-1]

            line_k = re.findall(r"\d+", line_n)

            info = {
                "レースタイトル": race_title,
                "競輪場":         race_stadium,
                "レース名":       race_name,
                "グレード":       race_gr,
                "開始時間":       race_start_time,
                "天気":           weather,
                "風速":           wind_speed,
                "レース番号":     race_info_header[0] if race_info_header else "",
                "開催日":         race_info_day,
                "開催番号":       race_info_header[3] if len(race_info_header) > 3 else "",
                "車立":           len(pd_racecard),
                "ライン数":       len(line_k),
                "ライン構成":     kousei2,
            }
        except Exception:
            pass

        return info

    # ──────────────────────────────────────────────────────
    #  内部: 正規化（パソ子AI 用DataFrame変換）
    # ──────────────────────────────────────────────────────

    def _normalize_entry(
        self,
        entry: pd.DataFrame,
        race_id: str,
    ) -> pd.DataFrame:
        """
        race_data_scrape.py の entry_table をパソ子AI が扱う
        標準DataFrameカラム形式に変換する
        """
        track_code  = race_id[8:10]
        race_no     = int(race_id[10:12])
        date_str    = f"{race_id[:4]}-{race_id[4:6]}-{race_id[6:8]}"
        bank_length = TRACK_BANK_LENGTH.get(track_code, 400)
        track_name  = TRACK_NAME_MASTER.get(track_code, f"場{track_code}")

        rows = []
        for _, r in entry.iterrows():
            # 落車・失格フラグ
            row_text = " ".join(str(v) for v in r.values)
            acc_flag = int(any(kw in row_text for kw in ["落車", "失格", "棄権", "妨害"]))

            # 着順（結果なし = 99）
            finish_pos = 99
            for col in ["着 順", "着順"]:
                if col in entry.columns:
                    try:
                        finish_pos = int(r[col])
                    except Exception:
                        pass
                    break

            # 番手（ライン内ポジション）
            line_pos = r.get("番手", 0)
            is_lead  = 1 if line_pos == 1 else 0

            rows.append({
                "race_id":        race_id,
                "race_date":      date_str,
                "track_name":     track_name,
                "track_code":     track_code,
                "race_number":    race_no,
                "grade":          self._normalize_grade(r.get("グレード", "")),
                "bank_length":    bank_length,
                "is_girls":       False,
                "is_challenge":   False,
                "car_number":     self._to_int(r.get("車 番", 0)),
                "frame_number":   self._to_int(r.get("枠 番", r.get("車 番", 0))),
                "player_name":    str(r.get("選手名", "")),
                "player_age":     self._to_int(r.get("年齢", 0)),
                "player_class":   self._to_int(r.get("期別", 0)),
                "player_grade":   str(r.get("級 班", "")),
                "race_score":     self._to_float(r.get("競走得点", 0)),
                "back_count":     self._to_int(r.get("B", 0)),
                "line_formation": str(r.get("ライン構成", "")),
                "line_code":      str(r.get("ライン", "")),
                "line_position":  self._to_int(line_pos),
                "is_line_leader": is_lead,
                "finish_position": finish_pos,
                "accident_flag":  acc_flag,
            })

        return pd.DataFrame(rows)

    def _normalize_grade(self, raw: str) -> str:
        """全角グレード → 英字グレードに変換"""
        mapping = {
            "ＧＰ": "GP", "Ｇ１": "G1", "Ｇ２": "G2", "Ｇ３": "G3",
            "Ｆ１": "F1", "Ｆ２": "F2",
        }
        return mapping.get(str(raw).strip(), str(raw).strip())

    # ──────────────────────────────────────────────────────
    #  内部: Excel保存
    # ──────────────────────────────────────────────────────

    def _save_excel(self, df: pd.DataFrame, path: str):
        """DataFrameをExcelに保存（ヘッダー色付き・オートフィット）"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            df.to_excel(path, index=False)

            wb  = load_workbook(path)
            ws  = wb.active
            hdr = PatternFill("solid", fgColor="1F4E79")
            fnt = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill      = hdr
                cell.font      = fnt
                cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = max(
                    (len(str(c.value)) if c.value else 0 for c in col), default=0
                )
                ws.column_dimensions[get_column_letter(col[0].column)].width = (
                    min(max_len + 2, 40)
                )

            ws.freeze_panes = "A2"
            wb.save(path)
        except Exception as e:
            logger.warning(f"Excel書式設定失敗（データは保存済み）: {e}")

    # ──────────────────────────────────────────────────────
    #  内部: HTTP ヘルパー
    # ──────────────────────────────────────────────────────

    def _check_scraping(self):
        if not _SCRAPING_AVAILABLE or self._session is None:
            raise RuntimeError(
                "requests / beautifulsoup4 がインストールされていません。\n"
                "pip install requests beautifulsoup4 lxml を実行してください。"
            )

    def _get_raw(self, url: str) -> Optional[str]:
        """リトライ付き GET → HTML文字列"""
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                resp = self._session.get(url, timeout=REQUEST_TIMEOUT)
                resp.raise_for_status()
                resp.encoding = resp.apparent_encoding
                return resp.text
            except Exception as e:
                if hasattr(e, "response") and e.response is not None:
                    if e.response.status_code == 404:
                        return None
                logger.debug(f"  [{attempt}/{MAX_RETRIES}] {e}  {url}")
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_INTERVAL)
        return None

    def _get_soup(self, url: str) -> Optional["BeautifulSoup"]:
        """リトライ付き GET → BeautifulSoup"""
        html = self._get_raw(url)
        return BeautifulSoup(html, "lxml") if html else None

    # ──────────────────────────────────────────────────────
    #  ユーティリティ
    # ──────────────────────────────────────────────────────

    @staticmethod
    def _to_float(v) -> float:
        try:
            return float(re.sub(r"[^\d.]", "", str(v)))
        except Exception:
            return 0.0

    @staticmethod
    def _to_int(v) -> int:
        try:
            return int(re.sub(r"[^\d]", "", str(v)))
        except Exception:
            return 0


# ===========================================================
# KeirinDataPreprocessor（変更なし）
# ===========================================================

class KeirinDataPreprocessor:
    """競輪データ前処理クラス"""

    def __init__(self, exclude_accidents: bool = True):
        self.exclude_accidents = exclude_accidents

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        データクリーニング
          - 落車・失格レースの除外
          - 欠損値処理
        """
        df_clean = df.copy()

        if self.exclude_accidents and "accident_flag" in df_clean.columns:
            before = len(df_clean)
            df_clean = df_clean[df_clean["accident_flag"] == 0]
            logger.info(f"落車・失格除外: {before - len(df_clean)}行除外 → {len(df_clean)}行")

        df_clean = df_clean.dropna(subset=["race_score", "back_count"])
        return df_clean

    def categorize_bank(self, bank_length: int) -> str:
        """バンク周長を3分類に変換"""
        if bank_length <= 335:
            return "BANK_333"
        elif bank_length <= 450:
            return "BANK_400"
        else:
            return "BANK_500"

    def categorize_race_type(
        self,
        num_racers: int,
        grade: str,
        is_girls: bool,
        is_challenge: bool,
    ) -> str:
        """レースタイプをカテゴリ分類"""
        if is_girls:
            return "GIRLS"
        if is_challenge:
            return "CHALLENGE"
        if grade in ("G1", "Ｇ１", "ＧＰ"):
            return "G1"
        if grade in ("G3", "Ｇ３"):
            return "G3"
        if num_racers == 7:
            return "RACE_7"
        if num_racers == 9:
            return "RACE_9"
        return "OTHER"

    def extract_line_formation(self, line_data: str) -> Dict[str, int]:
        """ライン構成を解析（例: '3-2-2' → {line_count: 3, ...}）"""
        if pd.isna(line_data) or line_data == "":
            return {"line_count": 0, "formation_type": "unknown"}
        try:
            parts     = [int(x) for x in str(line_data).split("-")]
            formation = {
                "line_count":    len(parts),
                "formation_type": f"{len(parts)}分戦",
            }
            for i, count in enumerate(parts, 1):
                formation[f"line{i}"] = count
            return formation
        except Exception:
            return {"line_count": 0, "formation_type": "error"}


# ===========================================================
# FeatureEngineering（変更なし）
# ===========================================================

class FeatureEngineering:
    """特徴量エンジニアリングクラス"""

    def create_features(self, df: pd.DataFrame) -> pd.DataFrame:
        """全特徴量をまとめて生成"""
        df = self.create_target_variable(df)
        df = self.create_flag4(df)
        df = self.create_ranking_features(df)
        return df

    def create_flag4(
        self, df: pd.DataFrame, race_id_col: str = "race_id"
    ) -> pd.DataFrame:
        """
        フラグ4: ライン先頭かつ競走得点・バック回数が共にレース内1位 → 1
        """
        df_r = df.copy()
        df_r["flag4"] = 0
        for _, group in df_r.groupby(race_id_col):
            max_score = group["race_score"].max()
            max_back  = group["back_count"].max()
            cond = (
                (group["is_line_leader"] == 1)
                & (group["race_score"] == max_score)
                & (group["back_count"] == max_back)
            )
            df_r.loc[group[cond].index, "flag4"] = 1
        return df_r

    def create_target_variable(
        self, df: pd.DataFrame, finish_position_col: str = "finish_position"
    ) -> pd.DataFrame:
        """目的変数: 3着以内 → 1, それ以外 → 0"""
        df_r = df.copy()
        df_r["target"] = (df_r[finish_position_col] <= 3).astype(int)
        return df_r

    def create_ranking_features(
        self, df: pd.DataFrame, race_id_col: str = "race_id"
    ) -> pd.DataFrame:
        """レース内ランキング特徴量（score_rank, back_rank）"""
        df_r = df.copy()
        for _, group in df_r.groupby(race_id_col):
            df_r.loc[group.index, "score_rank"] = group["race_score"].rank(
                ascending=False, method="min"
            )
            df_r.loc[group.index, "back_rank"] = group["back_count"].rank(
                ascending=False, method="min"
            )
        return df_r

    def prepare_features(
        self, df: pd.DataFrame, feature_columns: List[str]
    ) -> Tuple[pd.DataFrame, pd.Series]:
        """学習用特徴量とターゲットを返す"""
        X = df[feature_columns].copy()
        for col in X.select_dtypes(include=["object"]).columns:
            X[col] = pd.Categorical(X[col]).codes
        y = df["target"]
        return X, y


# ===========================================================
# サンプルデータ生成（テスト用、変更なし）
# ===========================================================

def create_sample_data(
    num_races: int = 100, num_racers_per_race: int = 9
) -> pd.DataFrame:
    """テスト用サンプルDataFrameを生成"""
    np.random.seed(42)
    data     = []
    race_id  = 1

    for _ in range(num_races):
        bank_length = np.random.choice([333, 400, 500])
        grade       = np.random.choice(["G3", "F1", "F2"], p=[0.1, 0.6, 0.3])
        is_girls    = np.random.choice([True, False], p=[0.1, 0.9])

        formations = (
            ["3-3-3", "3-4-2", "4-3-2", "5-2-2", "2-3-4"]
            if num_racers_per_race == 9
            else ["3-2-2", "3-3-1", "4-2-1"]
        )
        line_fmt = np.random.choice(formations)

        for car_num in range(1, num_racers_per_race + 1):
            data.append({
                "race_id":         race_id,
                "race_date":       f"2022-{np.random.randint(1,13):02d}-{np.random.randint(1,29):02d}",
                "track_name":      np.random.choice(["平塚", "川崎", "立川", "松戸"]),
                "track_code":      "51",
                "race_number":     np.random.randint(1, 13),
                "grade":           grade,
                "bank_length":     bank_length,
                "is_girls":        is_girls,
                "is_challenge":    False,
                "car_number":      car_num,
                "frame_number":    car_num,
                "player_id":       f"P{np.random.randint(1000, 9999)}",
                "player_name":     f"選手{car_num}",
                "player_age":      np.random.randint(18, 45),
                "player_class":    np.random.randint(70, 110),
                "player_grade":    np.random.choice(["S1", "S2", "A1", "A2"]),
                "race_score":      round(float(np.random.uniform(40, 70)), 2),
                "back_count":      int(np.random.randint(0, 5)),
                "line_formation":  line_fmt,
                "line_code":       str(np.random.randint(1, 4)),
                "line_position":   np.random.randint(0, 4),
                "is_line_leader":  1 if car_num in [1, 4, 7] else 0,
                "finish_position": np.random.randint(1, num_racers_per_race + 1),
                "accident_flag":   0,
                "race_title":      "テストレース",
                "race_start_time": "14:00",
                "weather":         "晴",
                "wind_speed":      "2",
            })
        race_id += 1

    return pd.DataFrame(data)


# ===========================================================
# __main__  テスト実行
# ===========================================================

if __name__ == "__main__":
    print("=" * 60)
    print(" data_processing.py v3.0 テスト")
    print("=" * 60)

    # ① サンプルデータでパイプラインテスト
    print("\n[1] サンプルデータ生成・前処理テスト")
    df = create_sample_data(num_races=50, num_racers_per_race=9)
    print(f"  生成: {len(df)}行 / {df['race_id'].nunique()}レース")

    pre = KeirinDataPreprocessor()
    df  = pre.clean_data(df)
    df["bank_category"] = df["bank_length"].apply(pre.categorize_bank)
    df["race_category"] = df.apply(
        lambda x: pre.categorize_race_type(9, x["grade"], x["is_girls"], x["is_challenge"]),
        axis=1,
    )

    fe = FeatureEngineering()
    df = fe.create_features(df)

    print(f"  カラム: {df.columns.tolist()}")
    print(f"  target 分布: {df['target'].value_counts().to_dict()}")
    print(f"  flag4  分布: {df['flag4'].value_counts().to_dict()}")

    # ② KeirinDataCollector の初期化確認
    print("\n[2] KeirinDataCollector 初期化テスト")
    col = KeirinDataCollector(
        start_year=2018,
        end_year=2025,
        output_dir="./scraped_data",
        resume=True,
    )
    print(f"  start_year={col.start_year}, end_year={col.end_year}")
    print(f"  スクレイピング可能: {_SCRAPING_AVAILABLE}")

    if _SCRAPING_AVAILABLE:
        print("\n[3] 1日分レースID取得テスト（2024-12-01）")
        ids = col.get_race_ids_for_date("2024-12-01")
        print(f"  取得レース数: {len(ids)}")
        for rid in ids[:5]:
            print(f"    {rid}")

        if ids:
            print(f"\n[4] 1レース スクレイピングテスト ({ids[0]})")
            df_one = col.scrape_one_race(ids[0])
            print(f"  取得行数: {len(df_one)}")
            print(df_one.head(3).to_string())
    else:
        print("\n  ⚠ requests/beautifulsoup4 未インストールのためスキップ")
        print("  インストール: pip install requests beautifulsoup4 lxml")

    print("\n=== テスト完了 ===")
