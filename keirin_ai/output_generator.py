"""
出力・可視化モジュール
予測結果をExcel出力し、ゾーンに応じたカラーリングを行う
"""

import pandas as pd
import numpy as np
from typing import Dict, List
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelOutputGenerator:
    """Excel出力生成クラス"""
    
    def __init__(self, output_dir: str = './output'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        # ゾーンカラー定義
        self.zone_colors = {
            'GACHI': 'FF0066CC',    # 濃い青
            'BLUE': 'FF3399FF',     # 青
            'TWILIGHT': 'FFFFCC66', # オレンジ
            'RED': 'FFFF3333'       # 赤
        }
        
        self.zone_names = {
            'GACHI': 'ガチゾーン',
            'BLUE': 'ブルーゾーン',
            'TWILIGHT': 'トワイライトゾーン',
            'RED': 'レッドゾーン'
        }
    
    def create_summary_excel(self, predictions: Dict, filename: str = 'predictions_summary.xlsx'):
        """
        予測サマリーExcelファイルを作成
        
        Args:
            predictions: {race_id: RacePrediction} の辞書
            filename: 出力ファイル名
        """
        filepath = os.path.join(self.output_dir, filename)
        
        # サマリーデータの作成
        summary_data = []
        for race_id, pred in predictions.items():
            race_info = pred.race_data.iloc[0]
            top3 = pred.get_top_players(3)
            
            summary_data.append({
                'レースID': race_id,
                'レース日': race_info.get('race_date', ''),
                '競輪場': race_info.get('track_name', ''),
                'レース番号': race_info.get('race_number', ''),
                'グレード': race_info.get('grade', ''),
                '◎': top3[top3['rank'] == 'A']['car_number'].iloc[0] if len(top3) >= 1 else '',
                '○': top3[top3['rank'] == 'B']['car_number'].iloc[0] if len(top3) >= 2 else '',
                '▲': top3[top3['rank'] == 'C']['car_number'].iloc[0] if len(top3) >= 3 else '',
                'A率': pred.a_rate,
                'CT値': pred.ct_value,
                'KS値': pred.ks_value,
                'ゾーン': self.zone_names[pred.zone],
                '推奨買い目': pred.get_recommendation()
            })
        
        df_summary = pd.DataFrame(summary_data)
        
        # Excelに出力
        df_summary.to_excel(filepath, index=False, sheet_name='予測サマリー')
        
        # カラーリング適用
        self._apply_zone_coloring(filepath, 'ゾーン')
        
        print(f"サマリーExcel作成完了: {filepath}")
        return filepath
    
    def create_detailed_excel(self, predictions: Dict, filename: str = 'predictions_detail.xlsx'):
        """
        詳細予測Excelファイルを作成（全選手の情報含む）
        
        Args:
            predictions: {race_id: RacePrediction} の辞書
            filename: 出力ファイル名
        """
        filepath = os.path.join(self.output_dir, filename)
        
        # 詳細データの作成
        detailed_data = []
        for race_id, pred in predictions.items():
            for idx, row in pred.race_data.iterrows():
                detailed_data.append({
                    'レースID': race_id,
                    'レース日': row.get('race_date', ''),
                    '競輪場': row.get('track_name', ''),
                    'レース番号': row.get('race_number', ''),
                    '車番': row.get('car_number', ''),
                    '選手名': row.get('player_name', ''),
                    '予想印': row['rank'],
                    '3着以内確率': row['probability'],
                    '予測': '◯' if row['prediction'] == 1 else '×',
                    '競走得点': row.get('race_score', ''),
                    'バック回数': row.get('back_count', ''),
                    'A率': pred.a_rate,
                    'CT値': pred.ct_value,
                    'KS値': pred.ks_value,
                    'ゾーン': self.zone_names[pred.zone]
                })
        
        df_detail = pd.DataFrame(detailed_data)
        
        # Excelに出力
        df_detail.to_excel(filepath, index=False, sheet_name='詳細予測')
        
        # カラーリング適用
        self._apply_zone_coloring(filepath, 'ゾーン')
        self._apply_rank_coloring(filepath, '予想印')
        
        print(f"詳細Excel作成完了: {filepath}")
        return filepath
    
    def _apply_zone_coloring(self, filepath: str, zone_column: str):
        """
        ゾーンに応じた行の背景色を適用
        
        Args:
            filepath: Excelファイルパス
            zone_column: ゾーン列名
        """
        wb = load_workbook(filepath)
        ws = wb.active
        
        # ヘッダー行を取得
        headers = [cell.value for cell in ws[1]]
        zone_col_idx = headers.index(zone_column) + 1 if zone_column in headers else None
        
        if zone_col_idx is None:
            print(f"警告: '{zone_column}'列が見つかりません")
            wb.save(filepath)
            return
        
        # 各行にカラーリング適用
        for row_idx in range(2, ws.max_row + 1):
            zone_value = ws.cell(row_idx, zone_col_idx).value
            
            # ゾーン名から対応するキーを取得
            zone_key = None
            for key, name in self.zone_names.items():
                if name == zone_value:
                    zone_key = key
                    break
            
            if zone_key and zone_key in self.zone_colors:
                fill_color = self.zone_colors[zone_key]
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                # 行全体に適用
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row_idx, col_idx).fill = fill
                    
                    # 文字色を調整（濃い色の場合は白文字に）
                    if zone_key in ['GACHI', 'RED']:
                        ws.cell(row_idx, col_idx).font = Font(color='FFFFFFFF', bold=True)
        
        wb.save(filepath)
    
    def _apply_rank_coloring(self, filepath: str, rank_column: str):
        """
        予想印（ランク）に応じたセルの背景色を適用
        
        Args:
            filepath: Excelファイルパス
            rank_column: ランク列名
        """
        wb = load_workbook(filepath)
        ws = wb.active
        
        # ヘッダー行を取得
        headers = [cell.value for cell in ws[1]]
        rank_col_idx = headers.index(rank_column) + 1 if rank_column in headers else None
        
        if rank_col_idx is None:
            wb.save(filepath)
            return
        
        # ランク別の色定義
        rank_colors = {
            'A': 'FFFF0000',  # 赤
            'B': 'FF0000FF',  # 青
            'C': 'FF00FF00',  # 緑
        }
        
        # 各行のランクセルにカラーリング
        for row_idx in range(2, ws.max_row + 1):
            rank_value = ws.cell(row_idx, rank_col_idx).value
            
            if rank_value in rank_colors:
                fill = PatternFill(
                    start_color=rank_colors[rank_value],
                    end_color=rank_colors[rank_value],
                    fill_type='solid'
                )
                ws.cell(row_idx, rank_col_idx).fill = fill
                ws.cell(row_idx, rank_col_idx).font = Font(color='FFFFFFFF', bold=True)
        
        wb.save(filepath)
    
    def create_statistics_excel(self, predictions: Dict, filename: str = 'statistics.xlsx'):
        """
        統計情報Excelファイルを作成
        
        Args:
            predictions: {race_id: RacePrediction} の辞書
            filename: 出力ファイル名
        """
        filepath = os.path.join(self.output_dir, filename)
        
        # ゾーン別統計
        zone_stats = {}
        for pred in predictions.values():
            zone = self.zone_names[pred.zone]
            if zone not in zone_stats:
                zone_stats[zone] = {'count': 0, 'a_rate_sum': 0, 'ct_value_sum': 0, 'ks_value_sum': 0}
            
            zone_stats[zone]['count'] += 1
            zone_stats[zone]['a_rate_sum'] += pred.a_rate
            zone_stats[zone]['ct_value_sum'] += pred.ct_value
            zone_stats[zone]['ks_value_sum'] += pred.ks_value
        
        # 統計DataFrameの作成
        stats_data = []
        for zone, stats in zone_stats.items():
            count = stats['count']
            stats_data.append({
                'ゾーン': zone,
                'レース数': count,
                '割合(%)': round(count / len(predictions) * 100, 2),
                '平均A率': round(stats['a_rate_sum'] / count, 4),
                '平均CT値': round(stats['ct_value_sum'] / count, 2),
                '平均KS値': round(stats['ks_value_sum'] / count, 4)
            })
        
        df_stats = pd.DataFrame(stats_data)
        
        # Excelに出力
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_stats.to_excel(writer, index=False, sheet_name='ゾーン別統計')
            
            # A率分布
            a_rates = [pred.a_rate for pred in predictions.values()]
            df_a_rate = pd.DataFrame({
                'A率範囲': ['0.9以上', '0.8-0.9', '0.7-0.8', '0.6-0.7', '0.6未満'],
                'レース数': [
                    sum(1 for x in a_rates if x >= 0.9),
                    sum(1 for x in a_rates if 0.8 <= x < 0.9),
                    sum(1 for x in a_rates if 0.7 <= x < 0.8),
                    sum(1 for x in a_rates if 0.6 <= x < 0.7),
                    sum(1 for x in a_rates if x < 0.6),
                ]
            })
            df_a_rate.to_excel(writer, index=False, sheet_name='A率分布')
        
        print(f"統計Excel作成完了: {filepath}")
        return filepath


class VisualizationGenerator:
    """可視化生成クラス"""
    
    def __init__(self, output_dir: str = './output'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def create_zone_distribution_chart(self, predictions: Dict, filename: str = 'zone_distribution.png'):
        """ゾーン分布グラフを作成"""
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            matplotlib.use('Agg')  # GUIなし環境用
            
            # 日本語フォント設定
            plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
            
            # ゾーン別カウント
            zone_names = {
                'GACHI': 'ガチゾーン',
                'BLUE': 'ブルーゾーン',
                'TWILIGHT': 'トワイライトゾーン',
                'RED': 'レッドゾーン'
            }
            
            zone_counts = {}
            for pred in predictions.values():
                zone = zone_names[pred.zone]
                zone_counts[zone] = zone_counts.get(zone, 0) + 1
            
            # グラフ作成
            plt.figure(figsize=(10, 6))
            colors = ['#0066CC', '#3399FF', '#FFCC66', '#FF3333']
            plt.bar(zone_counts.keys(), zone_counts.values(), color=colors)
            plt.title('Zone Distribution', fontsize=16)
            plt.xlabel('Zone', fontsize=12)
            plt.ylabel('Number of Races', fontsize=12)
            plt.grid(axis='y', alpha=0.3)
            
            filepath = os.path.join(self.output_dir, filename)
            plt.savefig(filepath, dpi=150, bbox_inches='tight')
            plt.close()
            
            print(f"グラフ作成完了: {filepath}")
            return filepath
        except Exception as e:
            print(f"グラフ作成エラー: {e}")
            return None


if __name__ == '__main__':
    # テスト実行
    print("=== 出力・可視化モジュールテスト ===")
    
    from prediction import RacePrediction
    
    # ダミー予測データ作成
    predictions = {}
    
    for i in range(5):
        race_data = pd.DataFrame({
            'race_id': [f'R{i+1:03d}'] * 9,
            'race_date': ['2022-12-15'] * 9,
            'track_name': ['平塚'] * 9,
            'race_number': [i+1] * 9,
            'grade': ['F1'] * 9,
            'car_number': list(range(1, 10)),
            'player_name': [f'選手{j}' for j in range(1, 10)],
            'race_score': np.random.uniform(40, 70, 9),
            'back_count': np.random.randint(0, 5, 9)
        })
        
        probabilities = np.random.uniform(0.1, 0.9, 9)
        probabilities = np.sort(probabilities)[::-1]
        
        pred = RacePrediction(f'R{i+1:03d}', race_data, probabilities)
        predictions[f'R{i+1:03d}'] = pred
    
    # Excel出力
    generator = ExcelOutputGenerator()
    generator.create_summary_excel(predictions)
    generator.create_detailed_excel(predictions)
    generator.create_statistics_excel(predictions)
    
    # グラフ生成
    viz = VisualizationGenerator()
    viz.create_zone_distribution_chart(predictions)
    
    print("\n出力完了!")
